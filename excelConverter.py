###############################
# Author: Will Miller         #
# Last Modified: 2/13/2020    #
# Version: 0.11               #
###############################

#openpyxl imports
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, numbers

#builtin imports
import datetime
import os
import time

#tkinter for display windows
from tkinter.filedialog import askopenfilename
import tkinter.messagebox


#### GLOBALS ####
NUMBER_FORMAT__CURRENCY = '"$"#,##0.00_-'
NUMBER_FORMAT__ACCOUNTING = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
NUMBER_FORMAT__DATE = "MM/DD/YYYY"
NUMBER_FORMAT__STANDARD = '#,##0.00'

BORDER__BOLD_UNDERLINE = Border(bottom=Side(border_style="thick"))
BORDER__FINAL_SUM = Border(top=Side(border_style="thick"), bottom=Side(border_style="double"))
BORDER__BOLD_ABOVELINE = Border(top=Side(border_style="thick"))

FONT__BOLD = Font(bold=True)
FONT__NORMAL = Font()

ALIGNMENT__WRAP_TEXT = Alignment(wrap_text=True)
ALIGNMENT__HORIZONAL_CENTER = Alignment(horizontal="center")
ALIGNMENT__HORIZONAL_LEFT = Alignment(horizontal="left")

COLUMN_WIDTH__DATE = 12
COLUMN_WIDTH__CURRENCY = 20
COLUMN_WIDTH__ROW_TITLE = 5
COLUMN_WIDTH__ROW_TITLE_SMALL = 3
COLUMN_WIDTH__EMPTY_ROW = 1

MIN_PAGE_WIDTH = 88
IDEAL_PAGE_WIDTH = 90
MAX_PAGE_WIDTH = 92
#### END GLOBALS ####

def print_debug(line):
    if 0:
        zprint(line)

def zprint(line):
    pass
    print(line)

class MigrateExcel:
    def __init__(self, inputWorkbookPath):
        self.iwb_path = inputWorkbookPath;

        #Derive output file based on input file
        self.owb_path = os.path.join(os.path.dirname(inputWorkbookPath), "modified_" + os.path.basename(inputWorkbookPath))
        zprint("Output workbook will be: \"%s\"" % (self.owb_path))

        #Initialize the pointers to the workbook variables to use later
        self.iwb = None
        self.owb = None
        self.iwb_do = None

        self.iwbSheetNameToActualName = dict()

        self.owbActualNameToSheetName = dict()

    #######################
    #  Starter Functions  #
    #######################
    def extractSheetNameMappings(self):
        #Build dictionary for mappings between iwb.sheetName -> actualName -> owb.sheetName
        for sheetName in self.iwb.sheetnames:
            currSheet = self.iwb[sheetName]
            actualName = currSheet["A2"].value

            self.iwbSheetNameToActualName[sheetName] = actualName

            if "Property on Hand at Beginning of Account" == actualName:
                self.owbActualNameToSheetName[actualName] = "Beginning"

            if "Property on Hand at Beginning of Account - Investment Detail" == actualName:
                self.owbActualNameToSheetName[actualName] = "Beginning Detail"

            if "Additional Property Received" == actualName:
                self.owbActualNameToSheetName[actualName] = "Additional"

            if "Schedule A - Receipts" == actualName:
                self.owbActualNameToSheetName[actualName] = "Sch A"

            if "Schedule B/E - For Export To Excel Only" == actualName:
                self.owbActualNameToSheetName[actualName] = "Sch B"

            if "Schedule C - Net Income from Trade or Business" == actualName:
                self.owbActualNameToSheetName[actualName] = "Sch C"

            if "Schedule D - Disbursements" == actualName:
                self.owbActualNameToSheetName[actualName] = "Sch D"

            if "Schedule F - Net Loss from Trade or Business" == actualName:
                self.owbActualNameToSheetName[actualName] = "Sch F"

            if "Schedule G - Distributions" == actualName:
                self.owbActualNameToSheetName[actualName] = "Sch G"

            if "Schedule H - Property on Hand at Close of Account" == actualName:
                self.owbActualNameToSheetName[actualName] = "Sch H"

            if "Schedule H - Investment Detail" == actualName:
                self.owbActualNameToSheetName[actualName] = "Sch H Detail"

            if "Estimated Market Value" == actualName:
                self.owbActualNameToSheetName[actualName] = "Market Value"

            if "Liability Detail" == actualName:
                self.owbActualNameToSheetName[actualName] = "Liability"

        zprint("\n### Mapping of input sheet to output sheet: ###")
        foundError = False
        for sheetName in self.iwb.sheetnames:
            try:
                zprint("%8s   --> %-65s --> %s" % (sheetName, self.iwbSheetNameToActualName[sheetName], self.owbActualNameToSheetName[self.iwbSheetNameToActualName[sheetName]]))
            except:
                zprint("ERROR: failed to map input sheet name: %s" % (sheetName))
                foundError = True

        if foundError:
            zprint("ERROR: Failed to map all input sheets! Cant continue.")
            exit(1)

        zprint("\n")

    def startMigration(self):
        self.unknownInputSheetNames = list()
        self.migrationFailedSheets = list()

        zprint("#################################")
        zprint("#   STARTING SHEET MIGRATIONS   #")
        zprint("#################################\n")

        #Kick off all sheet conversions one by one
        for iwbSheetName in self.iwb.sheetnames:
            #Get actual sheet name from input sheet name
            try:
                actualName = self.iwbSheetNameToActualName[iwbSheetName]
            except:
                zprint("ERROR: Unable to map input sheet name (%s) to actual sheet name" % (iwbSheetName))
                self.unknownInputSheetNames.append(iwbSheetName)
                continue

            #Get output sheet name from actual sheet name
            try:
                owbSheetName = self.owbActualNameToSheetName[actualName]
            except:
                zprint("ERROR: Unable to find actual sheet name mapping to output sheet mapping.")
                self.unknownInputSheetNames.append(iwbSheetName)
                continue

            migrationStatus = 0

            if owbSheetName == "Beginning":
                zprint("#Migrating Beginning")
                try:
                    migrationStatus = self.migrateBeginning(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % (owbSheetName))
                    zprint("   Fatal Error details: %s" % (str(e)))

            elif owbSheetName == "Beginning Detail":
                zprint("#Migrating Beginning Detail")
                try:
                    migrationStatus = self.migrateBeginningDetail(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % (owbSheetName))
                    zprint("   Fatal Error details: %s" % (str(e)))

            elif owbSheetName == "Additional":
                zprint("#Migrating Beginning Detail")
                try:
                    migrationStatus = self.migrateAdditional(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % (owbSheetName))
                    zprint("   Fatal Error details: %s" % (str(e)))

            elif owbSheetName == "Sch A":
                zprint("#Migrating Schedule A")
                try:
                    migrationStatus = self.migrateSchA(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % (owbSheetName))
                    zprint("   Fatal Error details: %s" % (str(e)))

            elif owbSheetName == "Sch B":
                zprint("#Migrating Schedule B")
                #Custom migration, Create output sheets B and E
                try:
                    migrationStatus = self.migrateSchB(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % (owbSheetName))
                    zprint("   Fatal Error details: %s" % (str(e)))

                zprint("#Migrating Schedule E")
                try:
                    migrationStatus = self.migrateSchB_E(iwbSheetName, "Sch E")
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % ("Sch E"))
                    zprint("   Fatal Error details: %s" % (str(e)))


            elif owbSheetName == "Sch C":
                zprint("#Migrating Schedule C")
                try:
                    migrationStatus = self.migrateSchC(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % ("Sch E"))
                    zprint("   Fatal Error details: %s" % (str(e)))

            elif owbSheetName == "Sch D":
                zprint("#Migrating Schedule D")
                try:
                    migrationStatus = self.migrateSchD(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % ("Sch E"))
                    zprint("   Fatal Error details: %s" % (str(e)))

            elif owbSheetName == "Sch F":
                zprint("#Migrating Schedule F")
                try:
                    migrationStatus = self.migrateSchF(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % ("Sch E"))
                    zprint("   Fatal Error details: %s" % (str(e)))

            elif owbSheetName == "Sch G":
                zprint("#Migrating Schedule G")
                try:
                    migrationStatus = self.migrateSchG(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % ("Sch E"))
                    zprint("   Fatal Error details: %s" % (str(e)))

            elif owbSheetName == "Sch H":
                zprint("#Migrating Schedule H")
                try:
                    migrationStatus = self.migrateSchH(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % ("Sch E"))
                    zprint("   Fatal Error details: %s" % (str(e)))

            elif owbSheetName == "Sch H Detail":
                zprint("#Migrating Schedule H Detail")
                try:
                    migrationStatus = self.migrateSchHDetail(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % ("Sch E"))
                    zprint("   Fatal Error details: %s" % (str(e)))

            elif owbSheetName == "Market Value":
                zprint("#Migrating Market Value")
                try:
                    migrationStatus = self.migrateMarketValue(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % ("Sch E"))
                    zprint("   Fatal Error details: %s" % (str(e)))

            elif owbSheetName == "Liability":
                zprint("#Migrating Liability")
                try:
                    migrationStatus = self.migrateLiability(iwbSheetName, owbSheetName)
                except Exception as e:
                    zprint("ERROR: Fatal error occurred while migrating %s sheet. Manual port will be required." % ("Sch E"))
                    zprint("   Fatal Error details: %s" % (str(e)))

            else:
                zprint("ERROR: Could not map input sheet (%s) to an ouptut sheet. Migration failed" % (iwbSheetName))
                self.unknownInputSheetNames.append(iwbSheetName)

            #Check if any sheet migrations failed
            if migrationStatus != 0:
                zprint("ERROR: Migration failed for sheet: %s (%s)\n" % (iwbSheetName, owbSheetName))
                self.migrationFailedSheets.append(iwbSheetName)

        #Create summary page details
        self.createSummaryPage()

        zprint("########################################")
        zprint("#   FINSHED STARTING SHEET MIGRATION   #")
        zprint("########################################\n")

        #Check if any input sheets were unknown
        if len(self.unknownInputSheetNames) > 0:
            zprint("ERROR: Found %d unkonown input sheets: %s" % (len(self.unknownInputSheetNames), ", ".join(self.unknownInputSheetNames)))

    def finalPolishing(self):
        print_debug("Opening workbook: %s" % (self.owb_path))

        #Overwrite pointer to self.owb to allow helper functions to still work here
        self.owb = load_workbook(self.owb_path)


        ##############################
        #    Fix Header Alignment    #
        ##############################
        for sheetName in self.owb.sheetnames:
            if sheetName == "Summary":
                continue
            currSheet = self.owb[sheetName]

            if sheetName == "Sch D":
                headerRow = 4
            else:
                headerRow = 5

            for row in currSheet.iter_rows(min_row=headerRow, max_row=headerRow):
                for cell in row:
                    cell.alignment = ALIGNMENT__HORIZONAL_CENTER

        ##############################
        #      Fix Column Widths     #
        ##############################
        zprint("\nFixing column widths for all sheets..")
        for sheetName in self.owb.sheetnames:
            currSheet = self.owb[sheetName]
            ##################################################
            #  Auto set width of each column based on data   #
            ##################################################
            dims = {}
            for row in currSheet.iter_rows(min_row=4, max_row=currSheet.max_row):
                for cell in row:
                    if cell.value:
                        if isinstance(cell.value, (int, float)):
                            # If its an int, increase its size
                            dims[cell.column_letter] = max(
                                (dims.get(cell.column_letter, 0), 1.8 * len(str(cell.value))))
                        else:
                            #If formula, ignore length
                            if isinstance(cell.value, str) and "=" in cell.value:
                                continue
                            # If its a string, pad it slightly
                            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))

            print_debug("\nAuto Align column width for Sheet: %s" % (sheetName))
            for col, value in dims.items():
                print_debug("  Col %s width %d" % (col, value))
                currSheet.column_dimensions[col].width = value

            #########################
            #  Custom Page Widths   #
            #########################
            if sheetName == "Summary":
                currSheet.column_dimensions["A"].width = 2
                currSheet.column_dimensions["B"].width = 16
                currSheet.column_dimensions["C"].width = 56
                currSheet.column_dimensions["D"].width = 16

            if sheetName == "Beginning":
                currSheet["A5"].alignment = ALIGNMENT__HORIZONAL_LEFT

                numOfTitleCols = 0
                for i in range(1, currSheet.max_column-2):
                    currSheet.column_dimensions[get_column_letter(i)].width = COLUMN_WIDTH__ROW_TITLE
                    numOfTitleCols += 1

                currSheet.column_dimensions["%c" % (get_column_letter(currSheet.max_column-2))].width = IDEAL_PAGE_WIDTH - ((numOfTitleCols * COLUMN_WIDTH__ROW_TITLE) + (2 * COLUMN_WIDTH__CURRENCY))
                currSheet.column_dimensions["%c" % (get_column_letter(currSheet.max_column-1))].width = COLUMN_WIDTH__CURRENCY
                currSheet.column_dimensions["%c" % (get_column_letter(currSheet.max_column))].width = COLUMN_WIDTH__CURRENCY

            if sheetName == "Beginning Detail":
                currSheet.column_dimensions["A"].width = COLUMN_WIDTH__ROW_TITLE_SMALL
                currSheet.column_dimensions["B"].width = 12
                currSheet.column_dimensions["C"].width = 40
                currSheet.column_dimensions["D"].width = 18
                currSheet.column_dimensions["E"].width = 18

            if sheetName == "Additional":
                currSheet.column_dimensions["A"].width = COLUMN_WIDTH__ROW_TITLE_SMALL
                currSheet.column_dimensions["B"].width = COLUMN_WIDTH__DATE
                currSheet.column_dimensions["C"].width = 22
                currSheet.column_dimensions["D"].width = 35
                currSheet.column_dimensions["E"].width = 18

            if sheetName == "Sch A":
                currSheet.column_dimensions["A"].width = COLUMN_WIDTH__EMPTY_ROW
                currSheet.column_dimensions["B"].width = 18
                currSheet.column_dimensions["C"].width = COLUMN_WIDTH__DATE
                currSheet.column_dimensions["D"].width = 30
                currSheet.column_dimensions["E"].width = 15
                currSheet.column_dimensions["F"].width = 15

            if sheetName == "Sch B":
                currSheet.column_dimensions["A"].width = COLUMN_WIDTH__ROW_TITLE_SMALL
                currSheet.column_dimensions["B"].width = COLUMN_WIDTH__DATE
                currSheet.column_dimensions["C"].width = 10
                currSheet.column_dimensions["D"].width = 23
                currSheet.column_dimensions["E"].width = 14
                currSheet.column_dimensions["F"].width = 14
                currSheet.column_dimensions["G"].width = 14

            if sheetName == "Sch E":
                currSheet.column_dimensions["A"].width = COLUMN_WIDTH__ROW_TITLE_SMALL
                currSheet.column_dimensions["B"].width = COLUMN_WIDTH__DATE
                currSheet.column_dimensions["C"].width = 10
                currSheet.column_dimensions["D"].width = 23
                currSheet.column_dimensions["E"].width = 14
                currSheet.column_dimensions["F"].width = 14
                currSheet.column_dimensions["G"].width = 14

            if sheetName == "Sch C":
                currSheet.column_dimensions["A"].width = COLUMN_WIDTH__ROW_TITLE_SMALL
                currSheet.column_dimensions["B"].width = COLUMN_WIDTH__ROW_TITLE_SMALL
                currSheet.column_dimensions["C"].width = COLUMN_WIDTH__DATE
                currSheet.column_dimensions["D"].width = 20
                currSheet.column_dimensions["E"].width = 30
                currSheet.column_dimensions["G"].width = 16

            if sheetName == "Sch D":
                dataHeaderRow = 4
                nameCol = self.getColNumByString(dataHeaderRow, "Name", owbSheetName=sheetName)
                for i in range(1,nameCol):
                    currSheet.column_dimensions["%c" % (get_column_letter(i))].width = COLUMN_WIDTH__EMPTY_ROW

                #Get rest of column numbers
                dateCol = nameCol + 1
                memoCol = nameCol + 2
                chkCol = nameCol + 3
                principalCol = nameCol + 4
                incomeCol = nameCol + 5

                currSheet.column_dimensions["%c" % (get_column_letter(nameCol))].width = 18
                currSheet.column_dimensions["%c" % (get_column_letter(dateCol))].width = COLUMN_WIDTH__DATE
                currSheet.column_dimensions["%c" % (get_column_letter(memoCol))].width = 22
                currSheet.column_dimensions["%c" % (get_column_letter(chkCol))].width = 6
                currSheet.column_dimensions["%c" % (get_column_letter(principalCol))].width = 14
                currSheet.column_dimensions["%c" % (get_column_letter(incomeCol))].width = 14

            if sheetName == "Sch F":
                currSheet.column_dimensions["A"].width = COLUMN_WIDTH__ROW_TITLE_SMALL
                currSheet.column_dimensions["B"].width = COLUMN_WIDTH__ROW_TITLE_SMALL
                currSheet.column_dimensions["C"].width = COLUMN_WIDTH__DATE
                currSheet.column_dimensions["D"].width = 20
                currSheet.column_dimensions["E"].width = 28
                currSheet.column_dimensions["F"].width = 8
                currSheet.column_dimensions["G"].width = 16

            if sheetName == "Sch G":
                currSheet.column_dimensions["A"].width = COLUMN_WIDTH__EMPTY_ROW
                currSheet.column_dimensions["B"].width = COLUMN_WIDTH__EMPTY_ROW
                currSheet.column_dimensions["C"].width = 16
                currSheet.column_dimensions["D"].width = COLUMN_WIDTH__DATE
                currSheet.column_dimensions["E"].width = 24
                currSheet.column_dimensions["F"].width = 8
                currSheet.column_dimensions["G"].width = 14
                currSheet.column_dimensions["H"].width = 14

            if sheetName == "Sch H":
                currSheet["A5"].alignment = ALIGNMENT__HORIZONAL_LEFT

                numOfTitleCols = 0
                for i in range(1, currSheet.max_column-2):
                    currSheet.column_dimensions[get_column_letter(i)].width = COLUMN_WIDTH__ROW_TITLE
                    numOfTitleCols += 1

                currSheet.column_dimensions["%c" % (get_column_letter(currSheet.max_column-2))].width = IDEAL_PAGE_WIDTH - ((numOfTitleCols * COLUMN_WIDTH__ROW_TITLE) + (2 * COLUMN_WIDTH__CURRENCY))
                currSheet.column_dimensions["%c" % (get_column_letter(currSheet.max_column-1))].width = COLUMN_WIDTH__CURRENCY
                currSheet.column_dimensions["%c" % (get_column_letter(currSheet.max_column))].width = COLUMN_WIDTH__CURRENCY

            if sheetName == "Sch H Detail":
                currSheet.column_dimensions["A"].width = COLUMN_WIDTH__ROW_TITLE_SMALL
                currSheet.column_dimensions["B"].width = 12
                currSheet.column_dimensions["C"].width = 40
                currSheet.column_dimensions["D"].width = 18
                currSheet.column_dimensions["E"].width = 18

            if sheetName == "Market Value":
                currSheet["A5"].alignment = ALIGNMENT__HORIZONAL_LEFT

                numOfTitleCols = 0
                for i in range(1, currSheet.max_column-2):
                    currSheet.column_dimensions[get_column_letter(i)].width = COLUMN_WIDTH__ROW_TITLE
                    numOfTitleCols += 1

                currSheet.column_dimensions["%c" % (get_column_letter(currSheet.max_column-2))].width = IDEAL_PAGE_WIDTH - ((numOfTitleCols * COLUMN_WIDTH__ROW_TITLE) + (2 * COLUMN_WIDTH__CURRENCY))
                currSheet.column_dimensions["%c" % (get_column_letter(currSheet.max_column-1))].width = COLUMN_WIDTH__CURRENCY
                currSheet.column_dimensions["%c" % (get_column_letter(currSheet.max_column))].width = COLUMN_WIDTH__CURRENCY

            if sheetName == "Liability":
                currSheet.column_dimensions["A"].width = COLUMN_WIDTH__EMPTY_ROW
                currSheet.column_dimensions["B"].width = COLUMN_WIDTH__EMPTY_ROW
                currSheet.column_dimensions["C"].width = COLUMN_WIDTH__EMPTY_ROW
                currSheet.column_dimensions["D"].width = 30
                currSheet.column_dimensions["E"].width = COLUMN_WIDTH__DATE
                currSheet.column_dimensions["F"].width = 10
                currSheet.column_dimensions["G"].width = 18
                currSheet.column_dimensions["H"].width = 18

        zprint("  Finished Fixing column widths for all sheets.")

        ##############################
        #     Check Page Width       #
        ##############################
        zprint("\nChecking page widths..")
        for sheetName in self.owb.sheetnames:
            currSheet = self.owb[sheetName]
            totalWidth = 0
            for i in range(1, currSheet.max_column+1):
                totalWidth += currSheet.column_dimensions[get_column_letter(i)].width

            if totalWidth < MIN_PAGE_WIDTH or totalWidth > MAX_PAGE_WIDTH:
                zprint("  WARNING: Page width of sheet %18s is %d. Desired page width: %d < pageWidth < %d" % (sheetName, totalWidth, MIN_PAGE_WIDTH, MAX_PAGE_WIDTH))

        #########################
        #    Fix Sheet Order    #
        #########################
        sheetOrder = ["Summary", "Beginning", "Beginning Detail", "Additional", "Sch A", "Sch B", "Sch C",
                      "Sch D", "Sch E", "Sch F", "Sch G", "Sch H", "Sch H Detail", "Market Value", "Liability"]

        print_debug("\nFixing Sheet order..")
        for i in range(0, len(self.owb._sheets)):
            currentSheetName = self.owb._sheets[i].title
            desiredSheetName = sheetOrder[i]
            if currentSheetName != desiredSheetName:
                #Find where desired sheet currently is and swap them
                indexOfDesiredSheet = -1
                for j in range(0, len(self.owb._sheets)):
                    sheetName = self.owb._sheets[j].title
                    if sheetName == desiredSheetName:
                        indexOfDesiredSheet = j
                        break

                #Swap sheets if it was found
                if indexOfDesiredSheet != -1:
                    print_debug("  Swapping sheet indexes %d <--> %d" % (i, j))
                    self.owb._sheets[i], self.owb._sheets[j] = self.owb._sheets[j], self.owb._sheets[i]

        zprint("\nDouble Checking that all %d sheets are in order.." % (len(self.owb._sheets)))
        numOutOfOrder = 0
        for i in range(0, len(self.owb._sheets)):
            currentSheetName = self.owb._sheets[i].title
            desiredSheetName = sheetOrder[i]
            if currentSheetName != desiredSheetName:
                numOutOfOrder += 1
                zprint("  WARNING: Sheet order does not match for sheet %s" % (currentSheetName))
                zprint("    Current = %18s -- Desired = %s" % (currentSheetName, desiredSheetName))
        if numOutOfOrder == 0:
            zprint("  All sheets are in order!")


        #########################
        #    Save new output    #
        #########################
        outputFilename = os.path.join(os.path.dirname(self.owb_path), "final_" + os.path.basename(self.owb_path))

        while os.path.exists(outputFilename):
            outputFilename = os.path.join(os.path.dirname(outputFilename), "final_" + os.path.basename(outputFilename))

        zprint("\n### COMPLETED MIGRATION! ###")
        zprint("   Final output sheet: %s" % (outputFilename))
        try:
            self.owb.save(filename=outputFilename)


        except:
            zprint("ERROR: Failed to save final output file (%s). Most likey because output file already exists" % (outputFilename))
            return -1
        else:
            print_debug("Removed intermediary file: %s" % (self.owb_path))
            os.remove(self.owb_path)

            directoryOfOutputFilename = os.path.dirname(outputFilename)
            os.startfile(directoryOfOutputFilename)

            #Open window printing that the conversion is complete
            #window = tkinter.Tk()
            #window.wm_withdraw()
            #tkinter.messagebox.showinfo(parent=window, title="Conversion Complete!", message="The conversion is complete. Final file can be found here:\n%s" % (outputFilename))



    ################################
    #     Workbook Functions       #
    ################################
    def openIWB_dataOnly(self):
        zprint("Opening input workbook as data only: %s" % (self.iwb_path))
        self.iwb_do = load_workbook(self.iwb_path, data_only=True)

    def openIWB(self):
        zprint("Opening input workbook: %s" % (self.iwb_path))
        self.iwb = load_workbook(self.iwb_path)

    def openOWB(self):
        self.owb = Workbook()

    def writeOWB(self):
        newOutPath = self.owb_path
        oldOutPath = newOutPath

        # Make sure output file name is unique, and do not overwrite old one.
        while os.path.exists(newOutPath):
            oldOutPath = newOutPath
            newOutPath = os.path.join(os.path.dirname(newOutPath), "new_" + os.path.basename(newOutPath))
            #zprint("WARNING: %s already exists. Renaming output file to: %s" % (oldOutPath, newOutPath))

        self.owb_path = newOutPath
        # Write workbook to new excel doc's unique name
        try:
            self.owb.save(filename=self.owb_path)
            print_debug("\nMigrated excel document created (no column width fixing yet): %s" % (self.owb_path))
        except:
            zprint("ERROR: Failed to write to output workbook: %s" % (self.owb_path))
            zprint("       -- Most likely due to workbook being open already.")


    ################################
    #       Helper Functions       #
    ################################
    def writeCell(self, sheet, cell, value, font=Font(), alignment=Alignment(), border=Border()):
        sheet[cell] = value
        sheet[cell].font = font
        sheet[cell].alignment = alignment
        sheet[cell].border = border

    def migratePageTitle(self, iwbSheetName, owbSheetName, titleColWidth, rowCount=3):
        print_debug("Migrating title for sheet: %10s to %s" % (iwbSheetName, owbSheetName))
        iwbCurrSheet = self.iwb[iwbSheetName]
        owbCurrSheet = self.owb[owbSheetName]

        title1 = iwbCurrSheet['A1'].value
        title2 = iwbCurrSheet['A2'].value
        date = iwbCurrSheet['A3'].value


        #Merge first %c columns of first three rows
        for i in range(1,rowCount+1):
            print_debug("  Column widths to merge for %s: A%d:%c%d" % (owbSheetName, i, titleColWidth, i))
            owbCurrSheet.merge_cells("A%d:%c%d" % (i, titleColWidth, i))

        #######################
        ## Write page header ##
        #######################
        for i in range(1, rowCount+1):
            if i == 1:
                self.writeCell(owbCurrSheet, "A%d" % (i), iwbCurrSheet['A%d' % (i)].value,
                               font=Font(bold=True, size=12), alignment=Alignment(horizontal="center"))
            if i == 2:
                self.writeCell(owbCurrSheet, "A%d" % (i), iwbCurrSheet['A%d' % (i)].value,
                               font=Font(bold=True, size=14), alignment=Alignment(horizontal="center"))
            if i == 3:
                self.writeCell(owbCurrSheet, "A%d" % (i), iwbCurrSheet['A%d' % (i)].value,
                               font=Font(bold=True, size=11), alignment=Alignment(horizontal="center"))

    def getRowRangeGeneric(self, iwbSheetName, col_letter, startString, endString):
        iwbCurrSheet = self.iwb[iwbSheetName]

        startRow = -1
        endRow = -1
        #Find Start/End of Current Assets
        for i in range(1, iwbCurrSheet.max_row+1):
            colData = iwbCurrSheet["%c%d" % (col_letter, i)].value
            if colData is None:
                continue

            if colData.lower() == startString.lower():
                startRow = i
            if colData.lower() == endString.lower():
                endRow = i

        if startRow == -1 or endRow == -1:
            zprint("ERORR: (getRowRangeGeneric) Unable to find start or end index for startString: %s, endString: %s on input page: %s (%s)" %
                  (startString, endString, iwbSheetName, self.iwbSheetNameToActualName[iwbSheetName]))
            return -1, -1
        else:
            return startRow, endRow
            print_debug("(getRowRangeGeneric) startString: %s, endString %s,    startRow: %s  -- endRow: %s" % (startString, endString, startRow, endRow))

    def findEmptyCols(self, owbSheetName):
        owbCurrSheet = self.owb[owbSheetName]
        emptyCols = list()

        colNum = 2
        for col in owbCurrSheet.iter_cols(min_col=colNum):
            isEmptyCol = True
            for cell in col:
                if cell.value is not None:
                    isEmptyCol = False
                    break
            if isEmptyCol:
                emptyCols.append(colNum)

            colNum += 1

        #Reverse list to aid in deletion
        emptyCols = emptyCols[::-1]

        return emptyCols

    def getRowNumByString(self, col_letter, searchValue, owbSheetName=None, iwbSheetName=None):
        if owbSheetName is not None and iwbSheetName is None:
            currSheet = self.owb[owbSheetName]
        elif owbSheetName is None and iwbSheetName is not None:
            currSheet = self.iwb[iwbSheetName]
        else:
            zprint("ERROR: Invalid use of this function!")
            return -1

        foundRow = -1
        for i in range(1, currSheet.max_row+1):
            colData = currSheet["%c%d" % (col_letter, i)].value
            if colData is None:
                continue

            try:
                if colData.lower() == searchValue.lower():
                    foundRow = i
            except:
                pass

        if foundRow == -1:
            zprint("  ERROR: Failed to find desired string: %s" % (searchValue))
            return -1
        else:
            return foundRow

    def getColNumByString(self, row_num, searchValue, owbSheetName=None, iwbSheetName=None ):
        if owbSheetName is not None and iwbSheetName is None:
            currSheet = self.owb[owbSheetName]
        elif owbSheetName is None and iwbSheetName is not None:
            currSheet = self.iwb[iwbSheetName]
        else:
            zprint("ERROR: Invalid use of this function!")
            return -1

        foundCol = -1
        for row in currSheet.iter_rows(min_row=row_num, max_row=row_num):
            colNum = 1
            for cell in row:
                #Skip empty cells
                if cell.value is None:
                    colNum += 1
                    continue

                #String match the value to the desired string
                try:
                    if searchValue.lower() == cell.value.lower():
                        foundCol = colNum
                except:
                    pass

                colNum += 1

        if foundCol == -1:
            zprint("  ERROR: Failed to find desired string: %s" % (searchValue))
            return -1
        else:
            return foundCol

    def dumbCopy(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb[owbSheetName]
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        for row in iwbCurrSheet.iter_rows(min_row=4):
            for cell in row:
                cellID = "%c%d" % (cell.column_letter, cell.row)
                #zprint("cellID: %s" % (cellID))
                #zprint("iwbValue: %s" % (cell.value))
                owbCurrSheet[cellID] = cell.value

    def dumbCopyWithRange(self, iwbSheetName, owbSheetName, startRow, endRow, keepFormulas=False):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb[owbSheetName]
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        for row in iwbCurrSheet.iter_rows(min_row=startRow, max_row=endRow):
            for cell in row:
                cellID = "%c%d" % (cell.column_letter, cell.row)
                #zprint("cellID: %s" % (cellID))
                #zprint("iwbValue: %s" % (cell.value))
                value = cell.value
                if isinstance(value, str) and "=" in value:
                    if keepFormulas:
                        owbCurrSheet[cellID].value = value
                    else:
                        owbCurrSheet[cellID].value = "FIX_FORMULA"
                        owbCurrSheet[cellID].font = Font(color='00FF0000')
                else:
                    owbCurrSheet[cellID].value = value

    def autoAlignColumnWidth(self, owbSheetName):
        owbCurrSheet = self.owb[owbSheetName]
        #Auto set width of each column based on data
        dims = {}
        for row in owbCurrSheet.iter_rows(min_row=4, max_row=owbCurrSheet.max_row):
            for cell in row:
                if cell.value:
                    if isinstance(cell.value, (int, float)):
                        #If its an int, increase its size
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), 1.8*len(str(cell.value))))
                    else:
                        #If its a string, pad it slightly
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), 5+len(str(cell.value))))

        print_debug("\nAuto Align column width for Sheet: %s" % (owbSheetName))
        for col, value in dims.items():
            print_debug("  Col %s width %d" % (col, value))
            owbCurrSheet.column_dimensions[col].width = value

    def getRowNumsForSubTotalRows(self, owbSheetName, colNumber, headerRowNum):
        owbCurrSheet = self.owb[owbSheetName]
        listOfTotalRows = list()

        #Create first entry in list as title header row for ease of total calculations
        listOfTotalRows.append((headerRowNum, "START"))

        priorCellValue = ""
        for col in owbCurrSheet.iter_cols(min_col=colNumber, max_col=colNumber):
            for cell in col:
                if cell.value is None:
                    priorCellValue = ""
                    continue

                #Find all rows that start with total
                if cell.value.lower().startswith("total"):
                    #Make sure prior row didnt have total in it as well.
                    #  If it did, the actual name of this section starts with total
                    if not priorCellValue.lower().startswith("total"):
                        #zprint("Found sub-total row: ", cell)
                        title = cell.value
                        rowNum = cell.row
                        #zprint("  Total: %s -- row Number: %d" % (title, rowNum))
                        listOfTotalRows.append((rowNum, title))
                    else:
                        print_debug("Skipping title %s because its not actually a total row." % (cell.value))

                priorCellValue = cell.value

        print_debug("  Found %d sub-total rows" % (len(listOfTotalRows)))

        return listOfTotalRows

    def migrateBeginning(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        #Migrate title, add 1 extra column for title merge due to extra column being added
        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth=get_column_letter(iwbCurrSheet.max_column+1))

        #Get data range for this sheet
        startRowOfAssets, endRowOfAssets = self.getRowRangeGeneric(iwbSheetName, "A", "assets", "total assets")
        if startRowOfAssets == -1 or endRowOfAssets == -1:
            return -1

        #Copy over original contents to start with
        self.dumbCopyWithRange(iwbSheetName, owbSheetName, startRowOfAssets, endRowOfAssets)

        #Insert new column for Carrying Value column
        owbCurrSheet.insert_cols(owbCurrSheet.max_column-1)

        #Get column letter for Carrying Value and Market Value columns
        carryingValueColLetter = get_column_letter(owbCurrSheet.max_column-2)
        marketValueColLetter = get_column_letter(owbCurrSheet.max_column-1)

        # Write new Value Titles
        self.writeCell(owbCurrSheet, "%c5" % (carryingValueColLetter), "Carrying Value",
                       font=Font(bold=True), border=Border(bottom=Side(border_style="thick")))
        self.writeCell(owbCurrSheet, "%c5" % (marketValueColLetter), "Market Value",
                       font=Font(bold=True), border=Border(bottom=Side(border_style="thick")))

        #Traverse data format to set bold and write out formulas
        startColNum = 1
        startRowNum = 5
        startCell = "%c%d" % (get_column_letter(startColNum), startRowNum)
        if owbCurrSheet[startCell].value != "ASSETS":
            zprint("ERROR: Failed to fix formulas and bolding because cell A5 did not contain \"ASSETS\". "
                   "Please fix this manually.")
            return -1
        else:
            #Find all column cells with "total" in them except overall total
            listOfTotalCells = list()
            for col in owbCurrSheet.iter_cols(min_col=1, max_col=owbCurrSheet.max_column-3):
                for cell in col:
                    if cell is not None and isinstance(cell.value, str):
                        #Hack to replace Checking/Savings with "Cash and Cash Equivalents
                        cell.value = cell.value.replace("Checking/Savings", "Cash and Cash Equivalents")

                        if cell.value.lower().startswith("total"):


                            totalCellRow = cell.row
                            totalCellCol = cell.column_letter

                            #Handle each total section
                            cell.font = FONT__BOLD

                            #Find start point of this total section
                            prevRow = totalCellRow - 1
                            foundStartOfTotalData = False

                            #Seach back through the rows of this column to find the start point. If not found, error out
                            while prevRow > 1:
                                #If cell has data, then its the start of the total section
                                if owbCurrSheet["%c%d" % (totalCellCol, prevRow)].value is not None:
                                    foundStartOfTotalData = True
                                    startOfTotalSectionRow = prevRow
                                    break
                                else:
                                    prevRow -= 1

                            #Make sure start of data was found
                            if not foundStartOfTotalData:
                                zprint("ERROR: Failed to find starting row for total cell: %s. Migration incomplete." % (cell.value))
                                return -1
                            else:
                                #Set section title to bold
                                owbCurrSheet["%c%d" % (totalCellCol, startOfTotalSectionRow)].font = FONT__BOLD

                                #Set formula
                                self.writeCell(owbCurrSheet, "%c%d" % (marketValueColLetter, totalCellRow), "FIX_ME",
                                               font=FONT__BOLD, border=BORDER__BOLD_ABOVELINE)

        #Set number format for money columns
        for row in range(startRowOfAssets+1, endRowOfAssets+1):
            owbCurrSheet["%c%d" % (carryingValueColLetter, row)].number_format = NUMBER_FORMAT__ACCOUNTING
            owbCurrSheet["%c%d" % (marketValueColLetter, row)].number_format = NUMBER_FORMAT__ACCOUNTING

        zprint("##Successfully migrated Beginning\n")
        return 0

    # Done
    def migrateBeginningDetail(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth="E")

        ########################
        ## Write table header ##
        ########################
        self.writeCell(owbCurrSheet, "B5", "QTY", font=Font(bold=True))
        self.writeCell(owbCurrSheet, "C5", "Investment", font=Font(bold=True))
        self.writeCell(owbCurrSheet, "D5", "Carrying Value", font=Font(bold=True))
        self.writeCell(owbCurrSheet, "E5", "Market Value", font=Font(bold=True))

        #Set cells to have thick bottom border
        owbCurrSheet["B5"].border = Border(bottom=Side(border_style="thick"))
        owbCurrSheet["C5"].border = Border(bottom=Side(border_style="thick"))
        owbCurrSheet["D5"].border = Border(bottom=Side(border_style="thick"))
        owbCurrSheet["E5"].border = Border(bottom=Side(border_style="thick"))

        ##############################################
        ##  Find row range of table for input data  ##
        ##############################################
        startRowOfData = -1
        endRowOfInventory = -1
        for i in range(1, iwbCurrSheet.max_row):
            colData = iwbCurrSheet["B%d" % (i)].value
            #zprint("colData: %s" % (colData))
            if colData is None:
                continue

            if colData.lower() == "inventory":
                startRowOfData = i+1
                print_debug("(migrateBeginningDetail) Inventory starts on row %d" % (startRowOfData))
            if colData.lower() == "total inventory":
                endRowOfInventory = i-1
                print_debug("(migrateBeginningDetail) End of inventory starts on row %d" % (endRowOfInventory))

        #Make sure data is extracted properly
        if startRowOfData == -1 or endRowOfInventory == -1:
            zprint("ERROR: (migrateBeginningDetail) Failed to extract start and end rows for data on sheet %s" % (iwbSheetName))
            return -1

        finalTotalRow = endRowOfInventory+1

        ##########################
        ## Write table contents ##
        ##########################
        #Get column letter for On Hand (Converts to QTY)
        onHandCol = self.getColNumByString(4, "On Hand", iwbSheetName=iwbSheetName)
        #Get column letter for Asset Value (Converts to Carrying Value)
        assetValueCol = self.getColNumByString(4, "Asset Value", iwbSheetName=iwbSheetName)

        if onHandCol == -1:
            zprint("ERROR: Unable to find On Hand column from input sheet %s." % (iwbSheetName))
            return -1

        if assetValueCol == -1:
            zprint("ERROR: Unable to find Asset Value column from input sheet %s." % (iwbSheetName))
            return -1

        for i in range(startRowOfData, finalTotalRow):
            #Copy investments
            owbCurrSheet["C%d" % i].value = iwbCurrSheet["C%d" % i].value

            #Copy On Hand -> QTY
            owbCurrSheet["B%d" % i].value = iwbCurrSheet["%c%d" % (get_column_letter(onHandCol), i)].value
            owbCurrSheet["B%d" % i].number_format = NUMBER_FORMAT__STANDARD

            #Copy Asset Value -> Carrying Value
            owbCurrSheet["D%d" % i].value = iwbCurrSheet["%c%d" % (get_column_letter(assetValueCol), i)].value
            #owbCurrSheet["D%d" % i].font = Font(color='00FF0000')
            owbCurrSheet["D%d" % i].number_format = NUMBER_FORMAT__ACCOUNTING

            #Copy Asset Value -> Market Value -- Write 0 instead
            #owbCurrSheet["E%d" % i].value = iwbCurrSheet["%c%d" % (get_column_letter(assetValueCol), i)].value
            owbCurrSheet["E%d" % i].value = 0
            owbCurrSheet["E%d" % i].number_format = NUMBER_FORMAT__ACCOUNTING


        ##########################
        ##   Write Final Row    ##
        ##########################
        self.writeCell(owbCurrSheet, "A%d" % finalTotalRow, "TOTAL", font=Font(bold=True))

        self.writeCell(owbCurrSheet, "D%d" % finalTotalRow, "=ROUND(SUM(D%d:D%d),5)" % (startRowOfData, endRowOfInventory), font=Font(bold=True))
        owbCurrSheet["D%d" % finalTotalRow].number_format = NUMBER_FORMAT__ACCOUNTING
        owbCurrSheet["D%d" % finalTotalRow].border = Border(top=Side(border_style="thick"), bottom=Side(border_style="double"))

        self.writeCell(owbCurrSheet, "E%d" % finalTotalRow, "=ROUND(SUM(E%d:E%d),5)" % (startRowOfData, endRowOfInventory), font=Font(bold=True))
        owbCurrSheet["E%d" % finalTotalRow].number_format = NUMBER_FORMAT__ACCOUNTING
        owbCurrSheet["E%d" % finalTotalRow].border = Border(top=Side(border_style="thick"), bottom=Side(border_style="double"))

        zprint("##Successfully migrated Beginning Detail\n")
        return 0

    # Completely Done
    def migrateAdditional(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth="E")

        #Check if page is empty -- if so, create empty formatted page
        rowThresholdToBeEmpty = 7
        if iwbCurrSheet.max_row < rowThresholdToBeEmpty:
            zprint("  INFO: Creating empty sheet for Additional page.")
            self.writeCell(owbCurrSheet, "B5", "Date", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "C5", "Name", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "D5", "Memo", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "E5", "Amount", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "A6", "TOTAL", font=FONT__BOLD)
            self.writeCell(owbCurrSheet, "E6", 0, font=FONT__BOLD, border=BORDER__FINAL_SUM)
            zprint("##Successfully migrated Additional\n")
            return 0

        #Dumb copy contents on page
        startRow, endRow = self.getRowRangeGeneric(iwbSheetName, "B", "Additional Property Received", "Total Additional Property Received")
        if startRow == -1 or endRow == -1:
            return -1

        #Copy data to new sheet
        self.dumbCopyWithRange(iwbSheetName, owbSheetName, startRow-1, endRow+1)

        ##########################
        #    Manipulate Rows     #
        ##########################
        #Delete row with row titles (Highest to lowest or else it will delete out of order
        owbCurrSheet.delete_rows(endRow)
        owbCurrSheet.delete_rows(startRow)

        #Insert extra row for data headers
        owbCurrSheet.insert_rows(4)

        ##########################
        #    Manipulate Cols     #
        ##########################
        #Delete empty columns
        for col in self.findEmptyCols(owbSheetName):
            owbCurrSheet.delete_cols(col)

        rowNumTOTAL = self.getRowNumByString("A", "TOTAL", owbSheetName=owbSheetName)
        if rowNumTOTAL == -1:
            return -1
        owbCurrSheet["A%d" % (rowNumTOTAL)].font = Font(bold=True)

        rowNumPaidAmount = self.getRowNumByString("E", "Paid Amount", owbSheetName=owbSheetName)
        if rowNumPaidAmount == -1:
            return -1

        #################################
        #  Manipulate Cell Formatting   #
        #################################
        #Set column titles to bold and underlined
        for row in owbCurrSheet.iter_rows(min_row=rowNumPaidAmount, max_row=rowNumPaidAmount):
            for cell in row:
                if cell.value is not None:
                    if cell.value == "Paid Amount":
                        cell.value = "Amount"
                    cell.font = Font(bold=True)
                    cell.border = Border(bottom=Side(border_style="thick"))

        #Set cell formatting
        for i in range(rowNumPaidAmount+1, rowNumTOTAL):
            owbCurrSheet["B%d" % (i)].number_format = NUMBER_FORMAT__DATE
            owbCurrSheet["E%d" % (i)].number_format = NUMBER_FORMAT__ACCOUNTING

        #Create Sum Formula
        if rowNumTOTAL != rowNumPaidAmount+1:
            print_debug("Formula: =ROUND(SUM(E%d:E%d),5)" % ((rowNumPaidAmount+1, rowNumTOTAL-1)))
            self.writeCell(owbCurrSheet, "E%d" % (rowNumTOTAL), "=ROUND(SUM(E%d:E%d),5)" % (rowNumPaidAmount+1, rowNumTOTAL-1),
                      border=BORDER__FINAL_SUM, font=FONT__BOLD)
            owbCurrSheet["E%d" % (rowNumTOTAL)].number_format = NUMBER_FORMAT__ACCOUNTING
        else:
            zprint("This sheet was empty. Not creating a final formula.")

        zprint("##Successfully migrated Additional\n")
        return 0

    # Completely Done
    def migrateSchA(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth="F")

        ###########################
        #  Get Data rows to copy  #
        ###########################
        startRow, endRow = 4, self.getRowNumByString("A", "total", iwbSheetName=iwbSheetName)
        if startRow == -1 or endRow == -1:
            return -1

        #Copy data to new sheet
        self.dumbCopyWithRange(iwbSheetName, owbSheetName, startRow, endRow+1)

        ##########################
        #    Manipulate Rows     #
        ##########################
        #Insert extra row for data headers
        owbCurrSheet.insert_rows(4)
        dataHeaderRow = 5

        ##########################
        #    Manipulate Cols     #
        ##########################
        #Delete empty columns
        for col in self.findEmptyCols(owbSheetName):
            owbCurrSheet.delete_cols(col)

        #Insert Principal Column
        owbCurrSheet.insert_cols(5)
        self.writeCell(owbCurrSheet, "%c%d" % (get_column_letter(5), dataHeaderRow), "Principal", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)

        #Add column header
        self.writeCell(owbCurrSheet, "B%d" % (dataHeaderRow), "Name", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)

        #################################
        #  Manipulate Cell Formatting   #
        #################################
        #Set row of column titles to bold and underlined
        for row in owbCurrSheet.iter_rows(min_row=dataHeaderRow, max_row=dataHeaderRow):
            for cell in row:
                if cell.value is not None:
                    if cell.value == "Paid Amount":
                        cell.value = "Income"
                    cell.font = Font(bold=True)
                    cell.border = Border(bottom=Side(border_style="thick"))

        #Get Final Total row
        finalTotalRowNum = self.getRowNumByString("A", "total", owbSheetName=owbSheetName)
        if finalTotalRowNum == -1:
            return -1

        NAME_COLUMN = "B"
        MEMO_COLUMN = "C"
        PRINCIPAL_COLUMN = "E"
        INCOME_COLUMN = "F"

        #Set cell formatting by column
        for i in range(dataHeaderRow+1, finalTotalRowNum+1):
            #Bold first column
            if owbCurrSheet["%c%d" % (NAME_COLUMN, i)].value is not None:
                if "total" not in owbCurrSheet["%c%d" % (NAME_COLUMN, i)].value.lower():
                    owbCurrSheet["%c%d" % (NAME_COLUMN, i)].font = FONT__BOLD

            #Set date column format
            owbCurrSheet["%c%d" % (MEMO_COLUMN, i)].number_format = NUMBER_FORMAT__DATE

            #Set Amount column format
            owbCurrSheet["%c%d" % (PRINCIPAL_COLUMN, i)].number_format = NUMBER_FORMAT__ACCOUNTING
            owbCurrSheet["%c%d" % (INCOME_COLUMN, i)].number_format = NUMBER_FORMAT__ACCOUNTING

        #################################
        #     Abbreviate Memo Names     #
        #################################
        memoCol = self.getColNumByString(dataHeaderRow, "Memo", owbSheetName=owbSheetName)
        if memoCol != -1:
            for col in owbCurrSheet.iter_cols(min_col=memoCol, max_col=memoCol):
                for cell in col:
                    if cell.value is not None and isinstance(cell.value, str):
                        if "Dividends received" in cell.value:
                            cell.value = "Dividends received"


        #################################
        #     Write Total Formulas      #
        #################################
        #Get list of sub-total rows
        listOfSubTotalRows = dict()
        listOfSubTotalRows = self.getRowNumsForSubTotalRows(owbSheetName, 2, dataHeaderRow)
        if len(listOfSubTotalRows) == 0:
            zprint("ERROR: Failed to extract row numbers for sub-total lines")
            return -1

        #Specify which column has the formulas for totals
        formulasColLetter = "F"

        #List starts at 1 because first entry always contains row number of column headers
        for i in range(1, len(listOfSubTotalRows)):
            subTotalStartRowNum, subTotalEndRowNum = listOfSubTotalRows[i-1][0], listOfSubTotalRows[i][0]

            #Add 2 because startRow is the prior total line
            #Subtract 1 because endRow is the total line
            sumRangeStart, sumRangeEnd = subTotalStartRowNum+2, subTotalEndRowNum-1

            #Write SUM formula to cell
            owbCurrSheet["%s%d" % (formulasColLetter, subTotalEndRowNum)].value =  "=SUM(%s%d:%s%d)" % (formulasColLetter, sumRangeStart, formulasColLetter, sumRangeEnd)
            owbCurrSheet["%s%d" % (formulasColLetter, subTotalEndRowNum)].border = BORDER__BOLD_ABOVELINE
            owbCurrSheet["%s%d" % (formulasColLetter, subTotalEndRowNum)].font = FONT__NORMAL

        ### Create final total formula: ###
        finalTotalRowFromSubTotals = listOfSubTotalRows[-1][0] + 1
        if finalTotalRowFromSubTotals != finalTotalRowNum:
            zprint("ERROR: Final Total row number does not match with expected. Error in alogrithm.")
            return -1

        #Create actual SUM formula
        finalSumFormula = "=SUM("
        for rowNum in [str(rowNum[0]) for rowNum in listOfSubTotalRows[1:]]:
            finalSumFormula += "%s%s" % (formulasColLetter, rowNum)
            finalSumFormula += "+"
        #Remove last + and replace with )
        finalSumFormula = finalSumFormula[:-1] + ")"

        #Write final total SUM formula
        print_debug("  Final Total: cell %s%d %s" % (formulasColLetter, finalTotalRowFromSubTotals, finalSumFormula))
        owbCurrSheet["%s%d" % (formulasColLetter, finalTotalRowFromSubTotals)].value = finalSumFormula
        owbCurrSheet["%s%d" % (formulasColLetter, finalTotalRowFromSubTotals)].border = BORDER__BOLD_ABOVELINE
        owbCurrSheet["%s%d" % (formulasColLetter, finalTotalRowFromSubTotals)].font = FONT__NORMAL

        #Add TOTAL PRINCIPAL AND INCOME after total
        self.writeCell(owbCurrSheet, "A%d" % (finalTotalRowNum+1), "TOTAL PRINCIPAL AND INCOME", font=FONT__BOLD)

        zprint("##Successfully migrated Schedule A\n")
        return 0

    # Splits into Sch B and Sch E sheets
    def migrateSchB(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth="G")
        owbCurrSheet["A2"].value = "Schedule B - Gains on Sales or Other Dispositions"

        ##########################
        #     Write Headers      #
        ##########################
        #Insert extra row for data headers
        owbCurrSheet.insert_rows(4)
        dataHeaderRow = 5

        self.writeCell(owbCurrSheet, "B%d" % (dataHeaderRow), "Date", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE, alignment=ALIGNMENT__HORIZONAL_CENTER)
        self.writeCell(owbCurrSheet, "C%d" % (dataHeaderRow), "Qty", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE, alignment=ALIGNMENT__HORIZONAL_CENTER)
        self.writeCell(owbCurrSheet, "D%d" % (dataHeaderRow), "Investment", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE, alignment=ALIGNMENT__HORIZONAL_CENTER)
        self.writeCell(owbCurrSheet, "E%d" % (dataHeaderRow), "Proceeds", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE, alignment=ALIGNMENT__HORIZONAL_CENTER)
        self.writeCell(owbCurrSheet, "F%d" % (dataHeaderRow), "Carrying Value", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE, alignment=ALIGNMENT__HORIZONAL_CENTER)
        self.writeCell(owbCurrSheet, "G%d" % (dataHeaderRow), "Gain", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE, alignment=ALIGNMENT__HORIZONAL_CENTER)

        ##########################
        #    Get IWB Columns     #
        ##########################
        #Get input WB columns for data
        dateColNum = self.getColNumByString(4, "Date", iwbSheetName=iwbSheetName)
        qtyColNum = self.getColNumByString(4, "Qty", iwbSheetName=iwbSheetName)
        itemColNum = self.getColNumByString(4, "Item", iwbSheetName=iwbSheetName)
        debitColNum = self.getColNumByString(4, "Debit", iwbSheetName=iwbSheetName)
        creditColNum = self.getColNumByString(4, "Credit", iwbSheetName=iwbSheetName)

        if dateColNum == -1 or qtyColNum == -1 or itemColNum == -1 or debitColNum == -1 or creditColNum == -1:
            zprint("  ERROR: Failed to get columns for Qty/Item/Debit/Credit from input workbook.")
            return -1

        ######################################
        #     Parse IWB and Copy to OWB      #
        ######################################
        #Begin parsing input WB and writing contents to output WB
        currentOWBRow = dataHeaderRow+1

        startRowForInputData = 6
        currIWBRow = startRowForInputData

        #Loop through all rows of input workbook
        while currIWBRow < iwbCurrSheet.max_row:
            itemRowOne = iwbCurrSheet["%s%d" % (get_column_letter(itemColNum), currIWBRow)].value
            itemRowTwo = iwbCurrSheet["%s%d" % (get_column_letter(itemColNum), currIWBRow+1)].value
            qtyValue = iwbCurrSheet["%s%d" % (get_column_letter(qtyColNum), currIWBRow)].value
            dateValue = iwbCurrSheet["%s%d" % (get_column_letter(dateColNum), currIWBRow)].value

            #Make sure the two rows are for the same item
            if itemRowOne != itemRowTwo:
                zprint("  ERROR: Consecutive rows didnt match for row %d! Skipping row to try and finish." % (currIWBRow))
                currIWBRow += 1
                continue

            try:
                creditValue = float(iwbCurrSheet["%s%d" % (get_column_letter(creditColNum), currIWBRow)].value)
            except:
                zprint("  ERROR: Failed to extract value from credit: ",
                      iwbCurrSheet["%s%d" % (get_column_letter(creditColNum), currIWBRow)].value)
                return -1

            try:
                debitValue = float(iwbCurrSheet["%s%d" % (get_column_letter(debitColNum), currIWBRow+1)].value)
            except:
                zprint("  ERROR: Failed to extract value from debit: ",
                      iwbCurrSheet["%s%d" % (get_column_letter(debitColNum), currIWBRow+1)].value)
                return -1

            #Check if it was a gain or a loss
            if creditValue >= debitValue:  #Gain
                print_debug("  Found Gain on rows %d and %d: %s" % (currIWBRow, currIWBRow+1, itemRowOne))

                #Write values to output workbook
                self.writeCell(owbCurrSheet, "B%d" % (currentOWBRow), dateValue)
                self.writeCell(owbCurrSheet, "C%d" % (currentOWBRow), qtyValue)
                self.writeCell(owbCurrSheet, "D%d" % (currentOWBRow), itemRowOne)
                self.writeCell(owbCurrSheet, "E%d" % (currentOWBRow), creditValue)
                self.writeCell(owbCurrSheet, "F%d" % (currentOWBRow), debitValue)
                self.writeCell(owbCurrSheet, "G%d" % (currentOWBRow), creditValue-debitValue)

                currentOWBRow += 1
            else: #loss
                print_debug("  Skipping rows %d and %d because its a loss. (Credit: %.2f, Debit: %.2f)" % (currIWBRow, currIWBRow+1, creditValue, debitValue))
                pass

            #Increment by two becuase input workbook has 2 lines per item
            currIWBRow += 2

        #################################
        #  Manipulate Cell Formatting   #
        #################################
        totalGainsRow = owbCurrSheet.max_row+2
        #Write TOTAL GAINS line
        self.writeCell(owbCurrSheet, "A%d" % (totalGainsRow), "TOTAL GAINS", font=FONT__BOLD)

        endRow = totalGainsRow

        #Set cell formatting by column
        for i in range(dataHeaderRow+1, endRow+2):
            #Set date column format
            owbCurrSheet["B%d" % (i)].number_format = NUMBER_FORMAT__DATE
            owbCurrSheet["C%d" % (i)].number_format = NUMBER_FORMAT__STANDARD

            owbCurrSheet["D%d" % (i)].alignment = ALIGNMENT__WRAP_TEXT

            #Set currency formats
            owbCurrSheet["E%d" % (i)].number_format = NUMBER_FORMAT__ACCOUNTING
            owbCurrSheet["F%d" % (i)].number_format = NUMBER_FORMAT__ACCOUNTING
            owbCurrSheet["G%d" % (i)].number_format = NUMBER_FORMAT__ACCOUNTING


        #######################
        #   Generate Totals   #
        #######################
        for colLetter in ["E", "F", "G"]:
            self.writeCell(owbCurrSheet, "%s%d" % (colLetter, endRow-1), "=ROUND(SUM(%s%d:%s%d),5)"
                           % (colLetter, dataHeaderRow+1, colLetter, endRow-2), border=BORDER__BOLD_ABOVELINE, font=FONT__BOLD)

        self.writeCell(owbCurrSheet, "G%d" % (endRow), "=G%d" % (endRow-1), border=BORDER__FINAL_SUM, font=FONT__BOLD)


        zprint("##Successfully migrated Schedule B\n")
        return 0

    # Splits into Sch B and Sch E sheets
    def migrateSchB_E(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth="G")
        owbCurrSheet["A2"].value = "Schedule E - Losses on Sales or Other Dispositions"

        ##########################
        #     Write Headers      #
        ##########################
        #Insert extra row for data headers
        owbCurrSheet.insert_rows(4)
        dataHeaderRow = 5

        self.writeCell(owbCurrSheet, "B%d" % (dataHeaderRow), "Date", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE, alignment=ALIGNMENT__HORIZONAL_CENTER)
        self.writeCell(owbCurrSheet, "C%d" % (dataHeaderRow), "Qty", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE, alignment=ALIGNMENT__HORIZONAL_CENTER)
        self.writeCell(owbCurrSheet, "D%d" % (dataHeaderRow), "Investment", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE, alignment=ALIGNMENT__HORIZONAL_CENTER)
        self.writeCell(owbCurrSheet, "E%d" % (dataHeaderRow), "Proceeds", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE, alignment=ALIGNMENT__HORIZONAL_CENTER)
        self.writeCell(owbCurrSheet, "F%d" % (dataHeaderRow), "Carrying Value", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE, alignment=ALIGNMENT__HORIZONAL_CENTER)
        self.writeCell(owbCurrSheet, "G%d" % (dataHeaderRow), "Loss", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE, alignment=ALIGNMENT__HORIZONAL_CENTER)

        ##########################
        #    Get IWB Columns     #
        ##########################
        #Get input WB columns for data
        dateColNum = self.getColNumByString(4, "Date", iwbSheetName=iwbSheetName)
        qtyColNum = self.getColNumByString(4, "Qty", iwbSheetName=iwbSheetName)
        itemColNum = self.getColNumByString(4, "Item", iwbSheetName=iwbSheetName)
        debitColNum = self.getColNumByString(4, "Debit", iwbSheetName=iwbSheetName)
        creditColNum = self.getColNumByString(4, "Credit", iwbSheetName=iwbSheetName)

        if dateColNum == -1 or qtyColNum == -1 or itemColNum == -1 or debitColNum == -1 or creditColNum == -1:
            zprint("  ERROR: Failed to get columns for Date/Qty/Item/Debit/Credit from input workbook.")
            return -1

        ######################################
        #     Parse IWB and Copy to OWB      #
        ######################################
        #Begin parsing input WB and writing contents to output WB
        currentOWBRow = dataHeaderRow+1

        startRowForInputData = 6
        currIWBRow = startRowForInputData

        #Loop through all rows of input workbook
        while currIWBRow < iwbCurrSheet.max_row:
            itemRowOne = iwbCurrSheet["%s%d" % (get_column_letter(itemColNum), currIWBRow)].value
            itemRowTwo = iwbCurrSheet["%s%d" % (get_column_letter(itemColNum), currIWBRow+1)].value
            qtyValue = iwbCurrSheet["%s%d" % (get_column_letter(qtyColNum), currIWBRow)].value
            dateValue = iwbCurrSheet["%s%d" % (get_column_letter(dateColNum), currIWBRow)].value

            #Make sure the two rows are for the same item
            if itemRowOne != itemRowTwo:
                zprint("  ERROR: Consecutive rows didnt match for row %d! Skipping row to try and finish." % (currIWBRow))
                currIWBRow += 1
                continue

            try:
                creditValue = float(iwbCurrSheet["%s%d" % (get_column_letter(creditColNum), currIWBRow)].value)
            except:
                zprint("  ERROR: Failed to extract value from credit: ",
                      iwbCurrSheet["%s%d" % (get_column_letter(creditColNum), currIWBRow)].value)
                return -1

            try:
                debitValue = float(iwbCurrSheet["%s%d" % (get_column_letter(debitColNum), currIWBRow+1)].value)
            except:
                zprint("  ERROR: Failed to extract value from debit: ",
                      iwbCurrSheet["%s%d" % (get_column_letter(debitColNum), currIWBRow+1)].value)
                return -1

            #Check if it was a gain or a loss
            if creditValue >= debitValue:  #Gain
                print_debug("  Skipping rows %d and %d because its a gain. (Credit: %.2f, Debit: %.2f)" % (currIWBRow, currIWBRow+1, creditValue, debitValue))
            else: #loss
                print_debug("  Found loss on rows %d and %d: %s" % (currIWBRow, currIWBRow+1, itemRowOne))
                #Write values to output workbook
                self.writeCell(owbCurrSheet, "B%d" % (currentOWBRow), dateValue)
                self.writeCell(owbCurrSheet, "C%d" % (currentOWBRow), qtyValue)
                self.writeCell(owbCurrSheet, "D%d" % (currentOWBRow), itemRowOne)
                self.writeCell(owbCurrSheet, "E%d" % (currentOWBRow), creditValue)
                self.writeCell(owbCurrSheet, "F%d" % (currentOWBRow), debitValue)
                self.writeCell(owbCurrSheet, "G%d" % (currentOWBRow), creditValue-debitValue)
                currentOWBRow += 1

            #Increment by two becuase input workbook has 2 lines per item
            currIWBRow += 2

        #################################
        #  Manipulate Cell Formatting   #
        #################################
        totalGainsRow = owbCurrSheet.max_row+2
        #Write TOTAL GAINS line
        self.writeCell(owbCurrSheet, "A%d" % (totalGainsRow), "TOTAL LOSSES", font=FONT__BOLD)

        endRow = totalGainsRow

        #Set cell formatting by column
        for i in range(dataHeaderRow+1, endRow+2):
            #Set date column format
            owbCurrSheet["B%d" % (i)].number_format = NUMBER_FORMAT__DATE
            owbCurrSheet["C%d" % (i)].number_format = NUMBER_FORMAT__STANDARD

            owbCurrSheet["D%d" % (i)].alignment = ALIGNMENT__WRAP_TEXT

            #Set currency formats
            owbCurrSheet["E%d" % (i)].number_format = NUMBER_FORMAT__ACCOUNTING
            owbCurrSheet["F%d" % (i)].number_format = NUMBER_FORMAT__ACCOUNTING
            owbCurrSheet["G%d" % (i)].number_format = NUMBER_FORMAT__ACCOUNTING


        #######################
        #   Generate Totals   #
        #######################
        for colLetter in ["E", "F", "G"]:
            self.writeCell(owbCurrSheet, "%s%d" % (colLetter, endRow-1), "=ROUND(SUM(%s%d:%s%d),5)"
                           % (colLetter, dataHeaderRow+1, colLetter, endRow-2), border=BORDER__BOLD_ABOVELINE, font=FONT__BOLD)

        self.writeCell(owbCurrSheet, "G%d" % (endRow), "=G%d" % (endRow-1), border=BORDER__FINAL_SUM, font=FONT__BOLD)


        zprint("##Successfully migrated Schedule E\n")
        return 0

    # - If net income is negative, Make empty sheet
    #  Competely done
    def migrateSchC(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]
        #Get sheet from input workbook with data only
        iwbCurrSheetDataOnly = self.iwb_do[iwbSheetName]

        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth="G")

        #Check if total is negative. If it is, leave sheet empty
        #totalRowNum = self.getRowNumByString("A", "Total", iwbSheetName=iwbSheetName)
        #balanceCol = self.getColNumByString(4, "Paid Amount", iwbSheetName=iwbSheetName)
        totalRowNum = iwbCurrSheet.max_row
        balanceCol = iwbCurrSheet.max_column
        if totalRowNum == -1 or balanceCol == -1:
            zprint("ERROR: Unable to find Total value on sheet: %s" % (iwbSheetName))
            return -1
        else:
            totalVal = iwbCurrSheetDataOnly["%c%d" % (get_column_letter(balanceCol), totalRowNum)].value
            print_debug("  Schedule C total: %s" % totalVal)
            try:
                totalVal = int(totalVal)
            except:
                totalVal = 0

            if totalVal <= 0:
                zprint("  Schedule C has a negative total. Creating empty sheet.")
                self.writeCell(owbCurrSheet, "C5", "Date", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
                self.writeCell(owbCurrSheet, "D5", "Name", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
                self.writeCell(owbCurrSheet, "E5", "Memo", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
                self.writeCell(owbCurrSheet, "F5", "Chk #", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
                self.writeCell(owbCurrSheet, "G5", "Amount", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
                self.writeCell(owbCurrSheet, "A8", "TOTAL", font=FONT__BOLD)
                self.writeCell(owbCurrSheet, "G8", 0, font=FONT__BOLD, border=BORDER__FINAL_SUM)
                owbCurrSheet["G8"].number_format = NUMBER_FORMAT__ACCOUNTING
                zprint("##Successfully migrated Schedule C\n")

                return 0
            else:
                #Schedule C has postiive total. Migrate data.
                pass

        ###########################
        #  Get Data rows to copy  #
        ###########################
        startRow, endRow = 4, self.getRowNumByString("A", "total", iwbSheetName=iwbSheetName)
        if startRow == -1 or endRow == -1:
            return -1

        #Copy data to new sheet
        self.dumbCopyWithRange(iwbSheetName, owbSheetName, startRow, endRow+1)

        ##########################
        #    Manipulate Rows     #
        ##########################
        #Insert extra row for data headers
        owbCurrSheet.insert_rows(4)
        dataHeaderRow = 5

        ##########################
        #    Manipulate Cols     #
        ##########################
        #Delete empty columns
        for col in self.findEmptyCols(owbSheetName):
            owbCurrSheet.delete_cols(col)

        #Delete columns with specified data header if they exist
        for colName in ["type", "balance"]:
            colToDel = self.getColNumByString(dataHeaderRow, colName, owbSheetName=owbSheetName)
            if colToDel == -1:
                zprint("  Didnt find column with header \"%s\" to delete. Skipping." % (colName))
                continue
            else:
                owbCurrSheet.delete_cols(colToDel)

        #################################
        #  Manipulate Cell Formatting   #
        #################################
        #Set row of column titles to bold and underlined
        for row in owbCurrSheet.iter_rows(min_row=dataHeaderRow, max_row=dataHeaderRow):
            for cell in row:
                if cell.value is not None:
                    if cell.value == "Paid Amount":
                        cell.value = "Amount"
                    if cell.value == "Num":
                        cell.value = "Chk #"
                    cell.font = Font(bold=True)
                    cell.border = Border(bottom=Side(border_style="thick"))


        endRow = self.getRowNumByString("A", "total", owbSheetName=owbSheetName)
        if endRow == -1:
            return -1

        #Set cell formatting by column
        for i in range(dataHeaderRow+1, endRow+1):
            #Bold first column
            if owbCurrSheet["B%d" % (i)].value is not None:
                if "total" not in owbCurrSheet["B%d" % (i)].value.lower():
                    owbCurrSheet["B%d" % (i)].font = FONT__BOLD

            #Set date column format
            owbCurrSheet["C%d" % (i)].number_format = NUMBER_FORMAT__DATE

            #Set Amount column format
            owbCurrSheet["G%d" % (i)].number_format = NUMBER_FORMAT__ACCOUNTING

            #Set Memo and Name text wrapping
            owbCurrSheet["D%d" % (i)].alignment = ALIGNMENT__WRAP_TEXT
            owbCurrSheet["E%d" % (i)].alignment = ALIGNMENT__WRAP_TEXT

        #################################
        #     Write Total Formulas      #
        #################################
        #Get Final Total row
        finalTotalRowNum = self.getRowNumByString("A", "total", owbSheetName=owbSheetName)
        if finalTotalRowNum == -1:
            return -1

        # Get list of sub-total rows
        listOfSubTotalRows = dict()
        listOfSubTotalRows = self.getRowNumsForSubTotalRows(owbSheetName, 2, dataHeaderRow)
        if len(listOfSubTotalRows) == 0:
            zprint("ERROR: Failed to extract row numbers for sub-total lines")
            return -1

        # Specify which column has the formulas for totals
        formulasColLetter = "G"

        # List starts at 1 because first entry always contains row number of column headers
        for i in range(1, len(listOfSubTotalRows)):
            subTotalStartRowNum, subTotalEndRowNum = listOfSubTotalRows[i - 1][0], listOfSubTotalRows[i][0]

            # Add 2 because startRow is the prior total line
            # Subtract 1 because endRow is the total line
            sumRangeStart, sumRangeEnd = subTotalStartRowNum + 2, subTotalEndRowNum - 1

            # Write SUM formula to cell
            owbCurrSheet["%s%d" % (formulasColLetter, subTotalEndRowNum)].value = "=SUM(%s%d:%s%d)" % (
            formulasColLetter, sumRangeStart, formulasColLetter, sumRangeEnd)
            owbCurrSheet["%s%d" % (formulasColLetter, subTotalEndRowNum)].border = BORDER__BOLD_ABOVELINE
            owbCurrSheet["%s%d" % (formulasColLetter, subTotalEndRowNum)].font = FONT__NORMAL

        ### Create final total formula: ###
        finalTotalRowFromSubTotals = listOfSubTotalRows[-1][0] + 1
        if finalTotalRowFromSubTotals != finalTotalRowNum:
            zprint("ERROR: Final Total row number does not match with expected. Error in alogrithm.")
            return -1

        # Create actual SUM formula
        finalSumFormula = "=SUM("
        for rowNum in [str(rowNum[0]) for rowNum in listOfSubTotalRows[1:]]:
            finalSumFormula += "%s%s" % (formulasColLetter, rowNum)
            finalSumFormula += "+"
        # Remove last + and replace with )
        finalSumFormula = finalSumFormula[:-1] + ")"

        # Write final total SUM formula
        print_debug("  Final Total: cell %s%d %s" % (formulasColLetter, finalTotalRowFromSubTotals, finalSumFormula))
        owbCurrSheet["%s%d" % (formulasColLetter, finalTotalRowFromSubTotals)].value = finalSumFormula
        owbCurrSheet["%s%d" % (formulasColLetter, finalTotalRowFromSubTotals)].border = BORDER__BOLD_ABOVELINE
        owbCurrSheet["%s%d" % (formulasColLetter, finalTotalRowFromSubTotals)].font = FONT__NORMAL

        zprint("##Successfully migrated Schedule C\n")
        return 0

    def migrateSchD(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        inputSheetDataHeaderRow = 4
        inputNameCol = self.getColNumByString(inputSheetDataHeaderRow, "Name", iwbSheetName=iwbSheetName)
        if inputNameCol == -1:
            zprint("  ERROR: Unable to find \"Name\" in input sheet header row. Failed.")
            return -1

        #Dynamically assign header column width based on Name column
        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth="%c" % (get_column_letter(inputNameCol+5)))

        #Get rows to copy
        startRow = 4
        endRow = self.getRowNumByString("A", "total", iwbSheetName=iwbSheetName)
        if startRow == -1 or endRow == -1:
            return -1

        #Copy data to new sheet
        self.dumbCopyWithRange(iwbSheetName, owbSheetName, startRow, endRow+1, keepFormulas=True)

        ##########################
        #    Manipulate Rows     #
        ##########################
        #NOTE: Dont insert extra row for this sheet to make fixing formulas easier
        dataHeaderRow = 4

        ##########################
        #    Manipulate Cols     #
        ##########################
        #Find Name column
        nameCol = self.getColNumByString(dataHeaderRow, "Name", owbSheetName=owbSheetName)
        if nameCol == -1:
            zprint("  ERROR: Unable to find \"Name\" column. Failed to port.")
            return -1

        #Delete empty columns after Name column
        for col in self.findEmptyCols(owbSheetName):
            if col > nameCol:
                owbCurrSheet.delete_cols(col)

        #Find Paid Amount column
        paidAmountCol = self.getColNumByString(dataHeaderRow, "Paid Amount", owbSheetName=owbSheetName)
        if paidAmountCol == -1:
            zprint("  ERROR: Unable to find \"Paid Amount\" column. Failed to port.")
            return -1

        #Add Income column
        self.writeCell(owbCurrSheet, "%c%d" % (get_column_letter(paidAmountCol+1), dataHeaderRow), "Income", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)

        ###################################
        #  Manipulate Header Formatting   #
        ###################################
        #Set row of column titles to bold and underlined
        for row in owbCurrSheet.iter_rows(min_row=dataHeaderRow, max_row=dataHeaderRow):
            for cell in row:
                if cell.value is not None:
                    #Rename data column titles
                    if cell.value == "Paid Amount":
                        cell.value = "Principal"
                    if cell.value == "Num":
                        cell.value = "Chk #"
                    cell.font = Font(bold=True)
                    cell.border = Border(bottom=Side(border_style="thick"))

        #Update endRow location
        endRow = self.getRowNumByString("A", "total", owbSheetName=owbSheetName)
        if endRow == -1:
            zprint("  ERROR: Failed to find row containing \"total\" in column A.")
            return -1

        #Add TOTAL PRINCIPAL AND INCOME after total
        self.writeCell(owbCurrSheet, "A%d" % (endRow+1), "TOTAL PRINCIPAL AND INCOME", font=FONT__BOLD)

        ###################################
        #  Set cell formatting by column  #
        ###################################
        nameCol = self.getColNumByString(dataHeaderRow, "Name", owbSheetName=owbSheetName)
        dateCol = self.getColNumByString(dataHeaderRow, "Date", owbSheetName=owbSheetName)
        memoCol = self.getColNumByString(dataHeaderRow, "Memo", owbSheetName=owbSheetName)
        principalCol = self.getColNumByString(dataHeaderRow, "Principal", owbSheetName=owbSheetName)

        if nameCol == -1 or dateCol == -1 or memoCol == -1 or principalCol == -1:
            zprint("  ERROR: Unable to find all columns (Name, Date, Memo, Principal) before applying formatting. Failed to port.")
            return -1

        for i in range(dataHeaderRow+1, endRow+2):
            #Bold header columns
            for j in range(2,nameCol):
                if owbCurrSheet["%c%d" % (get_column_letter(j),i)].value is not None:
                    #Fix alignment of all headers
                    owbCurrSheet["%c%d" % (get_column_letter(j), i)].alignment = None

                    #Bold all non-total headers
                    if "total" not in owbCurrSheet["%c%d" % (get_column_letter(j),i)].value.lower():
                        owbCurrSheet["%c%d" % (get_column_letter(j),i)].font = FONT__BOLD

            #Set Name column text wrap
            owbCurrSheet["%c%d" % (get_column_letter(nameCol), i)].alignment = ALIGNMENT__WRAP_TEXT
            #Set date column format
            owbCurrSheet["%c%d" % (get_column_letter(dateCol), i)].number_format = NUMBER_FORMAT__DATE
            #Set Memo column text wrap
            owbCurrSheet["%c%d" % (get_column_letter(memoCol), i)].alignment = ALIGNMENT__WRAP_TEXT
            #Set Amount column format
            owbCurrSheet["%c%d" % (get_column_letter(principalCol), i)].number_format = NUMBER_FORMAT__ACCOUNTING

        ##################
        #  Fix formulas  #
        ##################
        inputSheetAmountCol = self.getColNumByString(inputSheetDataHeaderRow, "Paid Amount", iwbSheetName=iwbSheetName)
        if inputSheetAmountCol == -1:
            zprint("  ERROR: Failed to find Paid Amount column from input sheet. Cannot fix formulas.")
            return -1

        #Fix formulas for principal (formerly Paid Amount) column
        for i in range(dataHeaderRow+1, endRow+2):
            currentCell = owbCurrSheet["%c%d" % (get_column_letter(principalCol), i)]
            if currentCell.value is not None and isinstance(currentCell.value, str):
                if "=" in currentCell.value:
                    #Prevent ROUND/SUM from being messed up by setting it to higher column letters
                    currentCell.value = currentCell.value.replace("ROUND", "XXX1")
                    currentCell.value = currentCell.value.replace("SUM", "YYY1")

                    #Fix formula
                    print_debug("Replacing %c with %c is formula: %s" % (get_column_letter(inputSheetAmountCol), get_column_letter(principalCol), currentCell.value))
                    currentCell.value = currentCell.value.replace("%c" % (get_column_letter(inputSheetAmountCol)), "%c" %(get_column_letter(principalCol)))

                    #Revert ROUND/SUM changes
                    currentCell.value = currentCell.value.replace("XXX1", "ROUND")
                    currentCell.value = currentCell.value.replace("YYY1", "SUM")

                    print_debug("New formula: %s" %(currentCell.value))

                    #Fix cell formatting as well
                    currentCell.border = BORDER__BOLD_ABOVELINE
                    currentCell.font = FONT__BOLD

        zprint("##Successfully migrated Schedule D\n")
        return 0

    # - If net income is positive, Make empty sheet
    # Done -- Totals arent tallied
    def migrateSchF(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]
        #Get sheet from input workbook with data only
        iwbCurrSheetDataOnly = self.iwb_do[iwbSheetName]

        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth="G")

        #Check if input sheet is empty
        rowThresholdToBeEmpty = 7
        if iwbCurrSheet.max_row < rowThresholdToBeEmpty:
            self.writeCell(owbCurrSheet, "C5", "Date", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "D5", "Name", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "E5", "Memo", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "F5", "Chk #", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "G5", "Amount", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "A8", "TOTAL", font=FONT__BOLD)
            self.writeCell(owbCurrSheet, "G8", 0, font=FONT__BOLD, border=BORDER__FINAL_SUM)
            owbCurrSheet["G8"].number_format = NUMBER_FORMAT__ACCOUNTING

            zprint("##Successfully migrated Schedule F\n")
            return 0

        try:

            #Check if total is positive. If it is, create sheet empty
            totalRowNum = self.getRowNumByString("A", "Total", iwbSheetName=iwbSheetName)
            balanceCol = self.getColNumByString(4, "Paid Amount", iwbSheetName=iwbSheetName)
            if totalRowNum == -1 or balanceCol == -1:
                zprint("ERROR: UNable to find Total value on sheet: %s" % (iwbSheetName))
                return -1
            else:
                #zprint("  Schedule F total: ", iwbCurrSheetDataOnly["%c%d" % (get_column_letter(balanceCol), totalRowNum)].value)
                if iwbCurrSheetDataOnly["%c%d" % (get_column_letter(balanceCol), totalRowNum)].value >= 0:
                    zprint("  Schedule F has a positive total. Creating empty sheet.")
                    self.writeCell(owbCurrSheet, "C5", "Date", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
                    self.writeCell(owbCurrSheet, "D5", "Name", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
                    self.writeCell(owbCurrSheet, "E5", "Memo", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
                    self.writeCell(owbCurrSheet, "F5", "Chk #", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
                    self.writeCell(owbCurrSheet, "G5", "Amount", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
                    self.writeCell(owbCurrSheet, "A8", "TOTAL", font=FONT__BOLD)
                    self.writeCell(owbCurrSheet, "G8", 0, font=FONT__BOLD, border=BORDER__FINAL_SUM)
                    owbCurrSheet["G8"].number_format = NUMBER_FORMAT__ACCOUNTING
                    zprint("##Successfully migrated Schedule F\n")
                    return 0
                else:
                    #Schedule C has postiive total. Migrate data.
                    pass

            ###########################
            #  Get Data rows to copy  #
            ###########################
            startRow, endRow = 4, self.getRowNumByString("A", "total", iwbSheetName=iwbSheetName)
            if startRow == -1 or endRow == -1:
                return -1

            #Copy data to new sheet
            self.dumbCopyWithRange(iwbSheetName, owbSheetName, startRow, endRow+1)


            ##########################
            #    Manipulate Rows     #
            ##########################
            #Insert extra row for data headers
            owbCurrSheet.insert_rows(4)
            dataHeaderRow = 5

            ##########################
            #    Manipulate Cols     #
            ##########################
            #Delete empty columns
            for col in self.findEmptyCols(owbSheetName):
                owbCurrSheet.delete_cols(col)

            #################################
            #  Manipulate Cell Formatting   #
            #################################
            #Set row of column titles to bold and underlined
            for row in owbCurrSheet.iter_rows(min_row=dataHeaderRow, max_row=dataHeaderRow):
                for cell in row:
                    if cell.value is not None:
                        if cell.value == "Paid Amount":
                            cell.value = "Amount"
                        if cell.value == "Num":
                            cell.value = "Chk #"
                        cell.font = Font(bold=True)
                        cell.border = Border(bottom=Side(border_style="thick"))

            #Get final row
            endRow = self.getRowNumByString("A", "total", owbSheetName=owbSheetName)

            if endRow == -1:
                return -1

            #Set cell formatting by column
            for i in range(dataHeaderRow+1, endRow+1):
                #Bold first column
                if owbCurrSheet["B%d" % (i)].value is not None:
                    if "total" not in owbCurrSheet["B%d" % (i)].value.lower():
                        owbCurrSheet["B%d" % (i)].font = FONT__BOLD

                #Set date column format
                owbCurrSheet["C%d" % (i)].number_format = NUMBER_FORMAT__DATE

                #Set Amount column format
                owbCurrSheet["G%d" % (i)].number_format = NUMBER_FORMAT__ACCOUNTING

            zprint("##Successfully migrated Schedule F\n")
            return 0
        except:
            return -1

    # ??
    def migrateSchG(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth="H")

        #Create empty sheet for now
        rowThresholdToBeEmpty = 7
        if iwbCurrSheet.max_row < rowThresholdToBeEmpty:
            self.writeCell(owbCurrSheet, "C5", "Name", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "D5", "Date", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "E5", "Memo", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "F5", "Chk #", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "G5", "Principal", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "H5", "Income", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "A8", "TOTAL", font=FONT__BOLD)
            self.writeCell(owbCurrSheet, "H8", 0, font=FONT__BOLD, border=BORDER__FINAL_SUM)
            owbCurrSheet["H8"].number_format = NUMBER_FORMAT__ACCOUNTING
            zprint(" INFO: Creating empty sheet for Schedule G")
        else:
            #Dumb copy for now
            self.dumbCopyWithRange(iwbSheetName, owbSheetName, 4, iwbCurrSheet.max_row, keepFormulas=True)

            #TODO: Fill in real code here when i know what to do
            pass

        zprint("##Successfully migrated Schedule G\n")

        return 0


    def migrateSchH(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        # Migrate title, add 1 extra column for title merge due to extra column being added
        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth=get_column_letter(iwbCurrSheet.max_column + 1))

        # Get data range for this sheet
        startRowOfAssets, endRowOfAssets = self.getRowRangeGeneric(iwbSheetName, "A", "assets", "total assets")
        if startRowOfAssets == -1 or endRowOfAssets == -1:
            return -1

        # Copy over original contents to start with
        self.dumbCopyWithRange(iwbSheetName, owbSheetName, startRowOfAssets, endRowOfAssets)

        # Insert new column for Carrying Value column
        owbCurrSheet.insert_cols(owbCurrSheet.max_column - 1)

        # Get column letter for Carrying Value and Market Value columns
        carryingValueColLetter = get_column_letter(owbCurrSheet.max_column - 2)
        marketValueColLetter = get_column_letter(owbCurrSheet.max_column - 1)

        # Write new Value Titles
        self.writeCell(owbCurrSheet, "%c5" % (carryingValueColLetter), "Carrying Value",
                       font=Font(bold=True), border=Border(bottom=Side(border_style="thick")))
        self.writeCell(owbCurrSheet, "%c5" % (marketValueColLetter), "Market Value",
                       font=Font(bold=True), border=Border(bottom=Side(border_style="thick")))

        # Traverse data format to set bold and write out formulas
        startColNum = 1
        startRowNum = 5
        startCell = "%c%d" % (get_column_letter(startColNum), startRowNum)
        if owbCurrSheet[startCell].value != "ASSETS":
            zprint("ERROR: Failed to fix formulas and bolding because cell A5 did not contain \"ASSETS\". "
                   "Please fix this manually.")
            return -1
        else:
            # Find all column cells with "total" in them except overall total
            listOfTotalCells = list()
            for col in owbCurrSheet.iter_cols(min_col=1, max_col=owbCurrSheet.max_column - 3):
                for cell in col:
                    if cell is not None and isinstance(cell.value, str):
                        #Hack to replace Checking/Savings with "Cash and Cash Equivalents
                        cell.value = cell.value.replace("Checking/Savings", "Cash and Cash Equivalents")

                        if cell.value.lower().startswith("total"):
                            totalCellRow = cell.row
                            totalCellCol = cell.column_letter

                            # Handle each total section
                            cell.font = FONT__BOLD

                            # Find start point of this total section
                            prevRow = totalCellRow - 1
                            foundStartOfTotalData = False

                            # Seach back through the rows of this column to find the start point. If not found, error out
                            while prevRow > 1:
                                # If cell has data, then its the start of the total section
                                if owbCurrSheet["%c%d" % (totalCellCol, prevRow)].value is not None:
                                    foundStartOfTotalData = True
                                    startOfTotalSectionRow = prevRow
                                    break
                                else:
                                    prevRow -= 1

                            # Make sure start of data was found
                            if not foundStartOfTotalData:
                                zprint(
                                    "ERROR: Failed to find starting row for total cell: %s. Migration incomplete." % (
                                        cell.value))
                                return -1
                            else:
                                # Set section title to bold
                                owbCurrSheet["%c%d" % (totalCellCol, startOfTotalSectionRow)].font = FONT__BOLD

                                # Set formula
                                self.writeCell(owbCurrSheet, "%c%d" % (marketValueColLetter, totalCellRow), "FIX_ME",
                                               font=FONT__BOLD, border=BORDER__BOLD_ABOVELINE)

        # Set number format for money columns
        for row in range(startRowOfAssets + 1, endRowOfAssets + 1):
            owbCurrSheet["%c%d" % (carryingValueColLetter, row)].number_format = NUMBER_FORMAT__ACCOUNTING
            owbCurrSheet["%c%d" % (marketValueColLetter, row)].number_format = NUMBER_FORMAT__ACCOUNTING

        zprint("##Successfully migrated Schedule H\n")
        return 0

    #Same as Beginning Detail code
    def migrateSchHDetail(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth="E")

        ########################
        ## Write table header ##
        ########################
        self.writeCell(owbCurrSheet, "B5", "QTY", font=Font(bold=True))
        self.writeCell(owbCurrSheet, "C5", "Investment", font=Font(bold=True))
        self.writeCell(owbCurrSheet, "D5", "Carrying Value", font=Font(bold=True))
        self.writeCell(owbCurrSheet, "E5", "Market Value", font=Font(bold=True))

        #Set cells to have thick bottom border
        owbCurrSheet["B5"].border = Border(bottom=Side(border_style="thick"))
        owbCurrSheet["C5"].border = Border(bottom=Side(border_style="thick"))
        owbCurrSheet["D5"].border = Border(bottom=Side(border_style="thick"))
        owbCurrSheet["E5"].border = Border(bottom=Side(border_style="thick"))

        ##############################################
        ##  Find row range of table for input data  ##
        ##############################################
        startRowOfData = -1
        endRowOfInventory = -1
        for i in range(1, iwbCurrSheet.max_row):
            colData = iwbCurrSheet["B%d" % (i)].value
            #zprint("colData: %s" % (colData))
            if colData is None:
                continue

            if colData.lower() == "inventory":
                startRowOfData = i+1
                print_debug("(migrateSchHDetail) Inventory starts on row %d" % (startRowOfData))
            if colData.lower() == "total inventory":
                endRowOfInventory = i-1
                print_debug("(migrateSchHDetail) End of inventory starts on row %d" % (endRowOfInventory))

        #Make sure data is extracted properly
        if startRowOfData == -1 or endRowOfInventory == -1:
            zprint("ERROR: (migrateSchHDetail) Failed to extract start and end rows for data on sheet %s" % (iwbSheetName))
            return -1

        finalTotalRow = endRowOfInventory+1

        ##########################
        ## Write table contents ##
        ##########################
        #Get column letter for On Hand (Converts to QTY)
        onHandCol = self.getColNumByString(4, "On Hand", iwbSheetName=iwbSheetName)
        #Get column letter for Asset Value (Converts to Carrying Value)
        assetValueCol = self.getColNumByString(4, "Asset Value", iwbSheetName=iwbSheetName)

        if onHandCol == -1:
            zprint("ERROR: Unable to find On Hand column from input sheet %s." % (iwbSheetName))
            return -1

        if assetValueCol == -1:
            zprint("ERROR: Unable to find Asset Value column from input sheet %s." % (iwbSheetName))
            return -1

        for i in range(startRowOfData, finalTotalRow):
            #Copy investments
            owbCurrSheet["C%d" % i].value = iwbCurrSheet["C%d" % i].value

            #Copy On Hand -> QTY
            owbCurrSheet["B%d" % i].value = iwbCurrSheet["%c%d" % (get_column_letter(onHandCol), i)].value
            owbCurrSheet["B%d" % i].number_format = NUMBER_FORMAT__STANDARD

            #Copy Asset Value -> Carrying Value
            owbCurrSheet["D%d" % i].value = iwbCurrSheet["%c%d" % (get_column_letter(assetValueCol), i)].value
            #owbCurrSheet["D%d" % i].font = Font(color='00FF0000')
            owbCurrSheet["D%d" % i].number_format = NUMBER_FORMAT__ACCOUNTING

            #Copy Asset Value -> Market Value -- Set value to 0 instead
            #owbCurrSheet["E%d" % i].value = iwbCurrSheet["%c%d" % (get_column_letter(assetValueCol), i)].value
            owbCurrSheet["E%d" % i].value = 0
            owbCurrSheet["E%d" % i].number_format = NUMBER_FORMAT__ACCOUNTING


        ##########################
        ##   Write Final Row    ##
        ##########################
        self.writeCell(owbCurrSheet, "A%d" % finalTotalRow, "TOTAL", font=Font(bold=True))

        self.writeCell(owbCurrSheet, "D%d" % finalTotalRow, "=ROUND(SUM(D%d:D%d),5)" % (startRowOfData, endRowOfInventory), font=Font(bold=True))
        owbCurrSheet["D%d" % finalTotalRow].number_format = NUMBER_FORMAT__ACCOUNTING
        owbCurrSheet["D%d" % finalTotalRow].border = Border(top=Side(border_style="thick"), bottom=Side(border_style="double"))

        self.writeCell(owbCurrSheet, "E%d" % finalTotalRow, "=ROUND(SUM(E%d:E%d),5)" % (startRowOfData, endRowOfInventory), font=Font(bold=True))
        owbCurrSheet["E%d" % finalTotalRow].number_format = NUMBER_FORMAT__ACCOUNTING
        owbCurrSheet["E%d" % finalTotalRow].border = Border(top=Side(border_style="thick"), bottom=Side(border_style="double"))

        zprint("##Successfully migrated Schedule H Detail\n")
        return 0

    def migrateMarketValue(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        #Migrate title
        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth=get_column_letter(iwbCurrSheet.max_column), rowCount=2)

        #Get data range for this sheet
        startRowOfAssets, endRowOfAssets = self.getRowRangeGeneric(iwbSheetName, "A", "assets", "total assets")
        if startRowOfAssets == -1 or endRowOfAssets == -1:
            return -1

        #Copy over original contents to start with
        self.dumbCopyWithRange(iwbSheetName, owbSheetName, startRowOfAssets-1, endRowOfAssets, keepFormulas=True)

        #Clear out title row contents
        owbCurrSheet.delete_rows(3)
        owbCurrSheet.insert_rows(3)

        #Get column letter for start and end dates
        startDateColLetter = get_column_letter(owbCurrSheet.max_column-1)
        endDateColLetter = get_column_letter(owbCurrSheet.max_column)

        #Traverse data format to set bold and write out formulas
        startColNum = 1
        startRowNum = 6
        startCell = "%c%d" % (get_column_letter(startColNum), startRowNum)
        if owbCurrSheet[startCell].value != "ASSETS":
            zprint("ERROR: Failed to fix formulas and bolding because cell A5 did not contain \"ASSETS\". "
                   "Please fix this manually.")
            return -1
        else:
            #Find all column cells with "total" in them except overall total
            listOfTotalCells = list()
            for col in owbCurrSheet.iter_cols(min_col=1, max_col=owbCurrSheet.max_column-3):
                for cell in col:
                    if cell is not None and isinstance(cell.value, str):
                        #Hack to replace Checking/Savings with "Cash and Cash Equivalents
                        cell.value = cell.value.replace("Checking/Savings", "Cash and Cash Equivalents")

                        if cell.value.lower().startswith("total"):
                            totalCellRow = cell.row
                            totalCellCol = cell.column_letter

                            #Handle each total section
                            cell.font = FONT__BOLD

                            #Find start point of this total section
                            prevRow = totalCellRow - 1
                            foundStartOfTotalData = False

                            #Seach back through the rows of this column to find the start point. If not found, error out
                            while prevRow > 1:
                                #If cell has data, then its the start of the total section
                                if owbCurrSheet["%c%d" % (totalCellCol, prevRow)].value is not None:
                                    foundStartOfTotalData = True
                                    startOfTotalSectionRow = prevRow
                                    break
                                else:
                                    prevRow -= 1

                            #Make sure start of data was found
                            if not foundStartOfTotalData:
                                zprint("ERROR: Failed to find starting row for total cell: %s. Migration incomplete." % (cell.value))
                                return -1
                            else:
                                #Set section title to bold
                                owbCurrSheet["%c%d" % (totalCellCol, startOfTotalSectionRow)].font = FONT__BOLD

                                #Set SUM cells to bold formatting
                                owbCurrSheet["%c%d" % (startDateColLetter, totalCellRow)].border = BORDER__BOLD_ABOVELINE
                                owbCurrSheet["%c%d" % (startDateColLetter, totalCellRow)].font = FONT__BOLD

                                owbCurrSheet["%c%d" % (endDateColLetter, totalCellRow)].border = BORDER__BOLD_ABOVELINE
                                owbCurrSheet["%c%d" % (endDateColLetter, totalCellRow)].font = FONT__BOLD

        #Fix Total Assets border formatting
        owbCurrSheet["%c%d" % (startDateColLetter, endRowOfAssets)].border = BORDER__FINAL_SUM
        owbCurrSheet["%c%d" % (endDateColLetter, endRowOfAssets)].border = BORDER__FINAL_SUM

        #Set number format for money columns
        for row in range(startRowOfAssets+1, endRowOfAssets+1):
            owbCurrSheet["%c%d" % (startDateColLetter, row)].number_format = NUMBER_FORMAT__ACCOUNTING
            owbCurrSheet["%c%d" % (endDateColLetter, row)].number_format = NUMBER_FORMAT__ACCOUNTING

        zprint("##Successfully migrated Market Value\n")
        return 0

    # Filled in dummy page -- ??
    def migrateLiability(self, iwbSheetName, owbSheetName):
        #Create new sheet in output workbook
        owbCurrSheet = self.owb.create_sheet(title=owbSheetName)
        #Get sheet from input workbook
        iwbCurrSheet = self.iwb[iwbSheetName]

        self.migratePageTitle(iwbSheetName, owbSheetName, titleColWidth="G")

        #Create empty sheet for now
        rowThresholdToBeEmpty = 6
        if iwbCurrSheet.max_row < rowThresholdToBeEmpty:
            zprint("  INFO: Creating empty sheet for Lability page beacuse row count is < %d" % (rowThresholdToBeEmpty))
            self.writeCell(owbCurrSheet, "C5", "Name", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "D5", "Date", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "E5", "Chk #", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "F5", "Amount", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "G5", "Balance", font=FONT__BOLD, border=BORDER__BOLD_UNDERLINE)
            self.writeCell(owbCurrSheet, "A8", "TOTAL", font=FONT__BOLD)
            self.writeCell(owbCurrSheet, "F8", 0, font=FONT__BOLD, border=BORDER__FINAL_SUM)
            self.writeCell(owbCurrSheet, "G8", 0, font=FONT__BOLD, border=BORDER__FINAL_SUM)
            owbCurrSheet["F8"].number_format = NUMBER_FORMAT__ACCOUNTING
            owbCurrSheet["G8"].number_format = NUMBER_FORMAT__ACCOUNTING
            owbCurrSheet.insert_cols(2)
        else:
            #Dumb copy for now
            self.dumbCopyWithRange(iwbSheetName, owbSheetName, 4, iwbCurrSheet.max_row, keepFormulas=True)

            # TODO: Fill in real code here when i know what to do
            pass

        zprint("##Successfully migrated Liability\n")

        return 0

    #Creates summary page, nothing else yet
    def createSummaryPage(self):
        owbCurrSheet = self.owb["Sheet"]
        owbCurrSheet.title = "Summary"

        #Find a sheet that has the original first line title
        for sheetName in self.iwb.sheetnames:
            iwbCurrSheet = self.iwb[sheetName]
            title1 = iwbCurrSheet['A1'].value
            if title1 is not None:
                break

        #Custom second line title for this page
        title2 = "Account Summary"

        #Get current year of this data
        date = iwbCurrSheet['A3'].value
        year = date.split()[-1]

        #Try and extract what year this came from
        try:
            year = int(year)
            if year > 2000 or year < 2100:
                date = "January %d through December %d" % (year-1, year)
        except:
            print("  WARNING: Unable to detect date range for summary page.")



        #Merge first %c columns of first three rows
        titleColWidth = 'D'
        for i in range(1,4):
            print_debug("  Column widths to merge for Summary page: A%d:%c%d" % (i, titleColWidth, i))
            owbCurrSheet.merge_cells("A%d:%c%d" % (i, titleColWidth, i))

        #Write first 3 title lines to page
        self.writeCell(owbCurrSheet, "A1", title1, font=Font(bold=True, size=12), alignment=Alignment(horizontal="center"))
        self.writeCell(owbCurrSheet, "A2", title2, font=Font(bold=True, size=16), alignment=Alignment(horizontal="center"))
        self.writeCell(owbCurrSheet, "A3", date, font=Font(bold=True, size=11), alignment=Alignment(horizontal="center"))

        #Write row section titles
        chargesRowNum, numOfChargeLines = 5, 5
        creditsRowNum, numOfCreditLines = 14, 5
        self.writeCell(owbCurrSheet, "B%d"%chargesRowNum, "CHARGES", font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))
        self.writeCell(owbCurrSheet, "B%d"%(chargesRowNum+numOfChargeLines+1), "TOTAL CHARGES:", font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))
        self.writeCell(owbCurrSheet, "B%d"%creditsRowNum, "CREDITS", font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))
        self.writeCell(owbCurrSheet, "B%d"%(creditsRowNum+numOfCreditLines+1), "TOTAL CREDITS:", font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))

        #Write Charges section contents
        self.writeCell(owbCurrSheet, "C%d"%(chargesRowNum+1), "Property on Hand at Beginning of Account (or Inventories)",
                  font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))
        self.writeCell(owbCurrSheet, "C%d"%(chargesRowNum+2), "Additional Property Received (or Supplemental Inventories)",
                  font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))
        self.writeCell(owbCurrSheet, "C%d"%(chargesRowNum+3), "Receipts (Schedule A)",
                  font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))
        self.writeCell(owbCurrSheet, "C%d"%(chargesRowNum+4), "Gains on Sales or Other Dispositions (Schedule B)",
                  font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))
        self.writeCell(owbCurrSheet, "C%d"%(chargesRowNum+5), "Net Income from Trade or Business (Schedule C)",
                  font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))

        #Write Credits section contents
        self.writeCell(owbCurrSheet, "C%d"%(creditsRowNum+1), "Disbursement (Schedule D)",
                  font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))
        self.writeCell(owbCurrSheet, "C%d"%(creditsRowNum+2), "Losses on Sales or Other Dispositions (Schedule E)",
                  font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))
        self.writeCell(owbCurrSheet, "C%d"%(creditsRowNum+3), "Net Loss from Trade or Business (Schedule F)",
                  font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))
        self.writeCell(owbCurrSheet, "C%d"%(creditsRowNum+4), "Distributions (Schedule G)",
                  font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))
        self.writeCell(owbCurrSheet, "C%d"%(creditsRowNum+5), "Property on Hand at Close of Account (Schedule H)",
                  font=Font(bold=True, size=11), alignment=Alignment(horizontal="left"))

        #TODO: Get values of each row from other pages

        #Create summary cell for Charges
        self.writeCell(owbCurrSheet, "D%d"%(chargesRowNum+numOfChargeLines+1), "=SUM(D%d:D%d)" % (chargesRowNum+1, chargesRowNum+numOfChargeLines))
        owbCurrSheet["D%d"%(chargesRowNum+numOfChargeLines+1)].border = Border(top=Side(border_style="thick"), bottom=Side(border_style="double"))
        owbCurrSheet["D%d"%(chargesRowNum+numOfChargeLines+1)].number_format = NUMBER_FORMAT__ACCOUNTING

        # Create summary cell for Credits
        self.writeCell(owbCurrSheet, "D%d"%(creditsRowNum+numOfCreditLines+1), "=SUM(D%d:D%d)" % (creditsRowNum+1, creditsRowNum+numOfCreditLines))
        owbCurrSheet["D%d"%(creditsRowNum+numOfCreditLines+1)].border = Border(top=Side(border_style="thick"), bottom=Side(border_style="double"))
        owbCurrSheet["D%d"%(creditsRowNum+numOfCreditLines+1)].number_format = NUMBER_FORMAT__ACCOUNTING


def main():
    #Get input file from user using Explorer
    validFile = False
    while not validFile:
        filename = askopenfilename()
        #Check if user cancelled or closed the file selection window
        if filename == "":
            zprint("No file was selected. Exiting.")
            exit(0)
        zprint("Excel file selected: %s" % (filename))

        #Make sure input file is an excel sheet
        name, ext = os.path.splitext(filename)
        zprint("File: %s -- Ext: %s" % (name, ext))

        if "xls" in ext:
            if os.path.exists(filename):
                validFile = True
            else:
                zprint("ERROR: File selected does not exist: \"%s\"" % (filename))
                validFile = False
        else:
            zprint("ERROR: You must provide an Excel document as input. The file you selected: %s" % (filename))
            validFile = False

    zprint("Converting file: %s" % (filename))

    #Create class to handle the migration
    migrateExcel = MigrateExcel(filename)

    #documetnation:
    # https://openpyxl.readthedocs.io/en/stable/usage.html

    #Open both input and output workbooks to start
    migrateExcel.openIWB()
    migrateExcel.openIWB_dataOnly()
    migrateExcel.openOWB()

    #Process input workbook to dynamically start tasks
    migrateExcel.extractSheetNameMappings()

    #Starts migrating all input workbook sheets to output workbook
    migrateExcel.startMigration()

    #Write final workbook
    migrateExcel.writeOWB()

    #Fix column widths, headers, and a few other cleanups
    # NOTE: This is required due to bug in openpyxl that wont allow column
    #       width assignments after writing values to cell
    migrateExcel.finalPolishing()

if __name__ == '__main__':
    main()
